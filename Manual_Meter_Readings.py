"""
Convertisseur de Relevés de Compteurs → JSON Meter-Readings
===========================================================
Application Streamlit mono-fichier

Auteur: Fabien
Version: 1.1.0 (avec support XML)
"""

# ============================================================================
# IMPORTS
# ============================================================================

import streamlit as st
import pandas as pd
import plotly.express as px
import json
import uuid
import re
import csv
import zipfile
import tempfile
import os
import io
import base64
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any
from dataclasses import dataclass, field
from io import StringIO
import openpyxl
import chardet

# ============================================================================
# SECTION 1 : CONSTANTES ET CONFIGURATION
# ============================================================================

# Sources disponibles (dropdown)
SOURCES = ["Amera", "Kamstrup", "Landis"]

# Préfixe par défaut pour mRID
DEFAULT_MRID_PREFIX = "ESR10307"

# Délimiteurs CSV supportés
CSV_DELIMITERS = [";", ",", "\t", "|"]
DEFAULT_DELIMITER = ";"

# Logo SVG intégré
LOGO_SVG = """<svg xmlns="http://www.w3.org/2000/svg" xml:space="preserve" class="logo-svg" viewBox="0 0 870.4 559.1"><style>.st0{fill:#fff}</style><g id="LOGO_AND_APP_ICON_LAYERS"><g id="XMLID_711_"><path id="XMLID_433_" d="M528 183.8c-2.9-5-8.4-8-14.1-8H384.2L429.8 92l-2.3 31.7c-6.8 4.3-11.3 11.9-11.3 20.5 0 13.4 10.9 24.2 24.2 24.2 13.4 0 24.2-10.9 24.2-24.2 0-7.1-3.1-13.5-8-17.9l6.6-91.2c.6-7.8-4.3-14.8-11.8-17s-15.3 1-19.1 7.9l-84.4 154.7c-2.8 5.1-2.7 11.3.3 16.3 2.9 5 8.4 8 14.1 8h129.6l-45.7 83.7 2.1-31.6c6.9-4.3 11.5-11.9 11.5-20.6 0-13.4-10.9-24.2-24.2-24.2-13.4 0-24.2 10.9-24.2 24.2 0 7 3 13.3 7.8 17.8l-6.1 90.8c-.5 7.8 4.3 14.7 11.8 16.9 1.6.5 3.1.7 4.7.7 5.9 0 11.4-3.2 14.4-8.6l84-154c3-5.1 2.9-11.4 0-16.3" class="st0"/><g id="XMLID_460_"><path id="XMLID_483_" d="M65.1 515q-18.75 0-31.5-8.4t-18-22.8l21.2-12.4c4.9 12.8 14.6 19.3 29 19.3 7 0 12.1-1.3 15.3-3.8s4.9-5.7 4.9-9.5c0-4.4-2-7.9-5.9-10.4-4-2.5-11-5.1-21.2-8-5.6-1.7-10.4-3.4-14.3-5-3.9-1.7-7.8-3.9-11.7-6.8-3.9-2.8-6.9-6.4-8.9-10.7s-3.1-9.4-3.1-15.1c0-11.4 4-20.5 12.1-27.3s17.8-10.2 29.2-10.2c10.2 0 19.2 2.5 26.9 7.5s13.8 11.9 18.1 20.8l-20.9 12.1c-5-10.8-13.1-16.2-24.1-16.2-5.2 0-9.2 1.2-12.1 3.5s-4.4 5.4-4.4 9.1c0 4 1.6 7.2 4.9 9.6 3.3 2.5 9.6 5.1 19 8 3.8 1.2 6.8 2.1 8.7 2.8 2 .7 4.7 1.7 8.1 3.1s6.1 2.7 7.9 3.9c1.9 1.2 4 2.8 6.4 4.9 2.4 2 4.2 4.1 5.5 6.3s2.3 4.8 3.2 7.8c.9 3.1 1.4 6.4 1.4 10 0 11.6-4.2 20.9-12.7 27.7Q85.5 515 65.1 515" class="st0"/><path id="XMLID_480_" d="M176.2 386.5c17.3 0 31.7 6.1 43.3 18.2s17.4 27.1 17.4 44.8c0 17.8-5.8 32.7-17.4 44.8s-26 18.2-43.3 18.2h-50v-126zm0 102.2c10.9 0 19.8-3.6 26.6-10.9s10.3-16.7 10.3-28.4c0-11.6-3.4-21.1-10.3-28.4-6.8-7.3-15.7-10.9-26.6-10.9H151v78.5h25.2z" class="st0"/><path id="XMLID_477_" d="M265.7 474.7c1.6 8 5.2 14.3 11.1 18.7 5.8 4.4 13 6.7 21.5 6.7 11.9 0 20.5-4.4 25.9-13.1l13.3 7.6c-8.8 13.6-22 20.3-39.6 20.3-14.3 0-25.9-4.5-34.8-13.4s-13.4-20.3-13.4-33.9 4.4-24.8 13.1-33.8c8.8-9 20.1-13.5 34-13.5 13.2 0 24 4.7 32.3 14s12.5 20.5 12.5 33.6c0 2.3-.2 4.6-.5 7h-75.4zm31.1-39.8c-8.4 0-15.4 2.4-20.9 7.1s-8.9 11.1-10.3 19h60.1c-1.3-8.5-4.7-15-10.1-19.4-5.3-4.5-11.6-6.7-18.8-6.7" class="st0"/><path id="XMLID_475_" d="M404.4 420.1c10.6 0 19 3.3 25.4 10s9.5 15.7 9.5 27.1v55.3h-15.7v-54.4c0-7.3-2-13-5.9-17-4-4-9.4-6-16.4-6-7.8 0-14.1 2.4-18.9 7.3s-7.2 12.3-7.2 22.4v47.7h-15.7v-90h15.7v13c6.4-10.3 16.1-15.4 29.2-15.4" class="st0"/><path id="XMLID_472_" d="M472.1 474.7c1.6 8 5.2 14.3 11.1 18.7 5.8 4.4 13 6.7 21.5 6.7 11.9 0 20.5-4.4 25.9-13.1l13.3 7.6c-8.8 13.6-22 20.3-39.6 20.3-14.3 0-25.9-4.5-34.8-13.4s-13.4-20.3-13.4-33.9 4.4-24.8 13.1-33.8c8.8-9 20.1-13.5 34-13.5 13.2 0 24 4.7 32.3 14s12.5 20.5 12.5 33.6c0 2.3-.2 4.6-.5 7h-75.4zm31.1-39.8c-8.4 0-15.4 2.4-20.9 7.1s-8.9 11.1-10.3 19h60.1c-1.3-8.5-4.7-15-10.1-19.4-5.3-4.5-11.6-6.7-18.8-6.7" class="st0"/><path id="XMLID_470_" d="M581.7 437.6c5.2-11.2 14.5-16.7 28.1-16.7v16.4c-7.7-.4-14.3 1.7-19.8 6.1s-8.3 11.6-8.3 21.4v47.7H566v-90h15.7z" class="st0"/><path id="XMLID_467_" d="M694.9 422.5h15.5v86.2c0 13.3-4.5 23.7-13.5 31q-13.5 11.1-32.4 11.1-15.15 0-26.1-5.7c-7.3-3.8-12.8-9.2-16.6-16.3l13.7-7.7c4.9 10.1 14.7 15.1 29.3 15.1 9.4 0 16.7-2.5 22-7.4s8-11.6 8-20.2v-12.8c-7.9 12-19.2 18-33.8 18-12.8 0-23.7-4.6-32.6-13.7s-13.3-20.2-13.3-33.3 4.4-24.1 13.3-33.2S648.1 420 661 420c14.8 0 26 5.9 33.8 17.8v-15.3zM640 489.8c6.1 6.1 13.7 9.2 22.9 9.2 9.1 0 16.7-3.1 22.9-9.2 6.1-6.1 9.2-13.7 9.2-22.9 0-9-3.1-16.6-9.2-22.7s-13.7-9.2-22.9-9.2c-9.1 0-16.7 3.1-22.9 9.2-6.1 6.1-9.2 13.7-9.2 22.7.1 9.1 3.1 16.8 9.2 22.9" class="st0"/><path id="XMLID_464_" d="M742.1 406.1c-3 0-5.5-1-7.6-3.1-2-2-3.1-4.5-3.1-7.4s1-5.4 3.1-7.5c2-2.1 4.6-3.1 7.6-3.1 2.9 0 5.3 1.1 7.4 3.1 2 2.1 3.1 4.6 3.1 7.5s-1 5.3-3.1 7.4-4.5 3.1-7.4 3.1m-7.9 106.4v-90h15.7v90z" class="st0"/><path id="XMLID_461_" d="M784.8 474.7c1.6 8 5.2 14.3 11.1 18.7 5.8 4.4 13 6.7 21.5 6.7 11.9 0 20.5-4.4 25.9-13.1l13.3 7.6c-8.8 13.6-22 20.3-39.6 20.3-14.3 0-25.9-4.5-34.8-13.4s-13.4-20.3-13.4-33.9 4.4-24.8 13.1-33.8c8.8-9 20.1-13.5 34-13.5 13.2 0 24 4.7 32.3 14s12.5 20.5 12.5 33.6c0 2.3-.2 4.6-.5 7h-75.4zm31.1-39.8c-8.4 0-15.4 2.4-20.9 7.1s-8.9 11.1-10.3 19h60.1c-1.3-8.5-4.7-15-10.1-19.4-5.3-4.5-11.6-6.7-18.8-6.7" class="st0"/></g></g></g></svg>"""

# Extensions acceptées
ACCEPTED_DATA_EXTENSIONS = [".csv", ".xlsx", ".xls", ".xml"]
ACCEPTED_ARCHIVE_EXTENSIONS = [".zip"]

# Profils de charge → (nom, intervalle)
LOAD_PROFILES = {
    "1-0:99.1.0": ("Profil de charge 1", "15min"),
    "1-0:99.2.0": ("Profil de charge 2", "24h")
}


def extract_lp_abbreviation(load_profile: str) -> str:
    """
    Extrait l'abréviation LP depuis le nom du profil de charge.

    Exemples:
        - "Profil de charge 1" → "LP1"
        - "Profil de charge 2" → "LP2"
        - "Load Profile 1" → "LP1"
        - "Valeurs de facturation" → "" (pas de LP)

    Args:
        load_profile: Nom du profil de charge

    Returns:
        Abréviation "LPX" ou chaîne vide si pas de profil numéroté
    """
    import re
    # Chercher un chiffre dans le nom du profil
    match = re.search(r'(\d+)', load_profile)
    if match and ('profil' in load_profile.lower() or 'load' in load_profile.lower()):
        return f"LP{match.group(1)}"
    return ""


# DST (Daylight Saving Time) → Offset timezone Europe/Zurich
# 0 = heure d'hiver (STD), 8 = heure d'été (DST)
DST_OFFSETS = {
    0: "+01:00",
    8: "+02:00"
}

# Unités nécessitant multiplication ×1000 (kWh → Wh, etc.)
UNITS_MULTIPLY_1000 = ["kwh", "kw", "kvarh", "kvah"]

# Mapping OBIS → IEC
# Clé: (code_obis, intervalle) → code_iec
OBIS_TO_IEC = {
    # Énergie active import (+A) - 1.8.x
    ("1-0:1.8.0", "15min"): "0.0.2.1.1.1.12.0.0.0.0.0.0.0.0.0.72.0",
    ("1-0:1.8.0", "24h"): "0.0.4.1.1.1.12.0.0.0.0.0.0.0.0.0.72.0",
    ("1-0:1.8.1", "24h"): "0.0.4.1.1.1.12.0.0.0.0.1.0.0.0.0.72.0",
    ("1-0:1.8.2", "24h"): "0.0.4.1.1.1.12.0.0.0.0.2.0.0.0.0.72.0",
    ("1-0:1.8.3", "24h"): "0.0.4.1.1.1.12.0.0.0.0.3.0.0.0.0.72.0",
    ("1-0:1.8.4", "24h"): "0.0.4.1.1.1.12.0.0.0.0.4.0.0.0.0.72.0",
    ("1-0:1.6.0", "NULL"): "0.8.0.0.1.1.37.0.0.0.0.0.0.0.0.0.38.0",
    
    # Énergie active export (-A) - 2.8.x
    ("1-0:2.8.0", "15min"): "0.0.2.1.19.1.12.0.0.0.0.0.0.0.0.0.72.0",
    ("1-0:2.8.0", "24h"): "0.0.4.1.19.1.12.0.0.0.0.0.0.0.0.0.72.0",
    ("1-0:2.8.1", "24h"): "0.0.4.1.19.1.12.0.0.0.0.1.0.0.0.0.72.0",
    ("1-0:2.8.2", "24h"): "0.0.4.1.19.1.12.0.0.0.0.2.0.0.0.0.72.0",
    ("1-0:2.8.3", "24h"): "0.0.4.1.19.1.12.0.0.0.0.3.0.0.0.0.72.0",
    ("1-0:2.8.4", "24h"): "0.0.4.1.19.1.12.0.0.0.0.4.0.0.0.0.72.0",
    ("1-0:2.6.0", "NULL"): "0.8.0.0.19.1.37.0.0.0.0.0.0.0.0.0.38.0",
    
    # Énergie réactive R1 (Ri+) - 5.8.x
    ("1-0:5.8.0", "15min"): "0.0.2.1.15.1.12.0.0.0.0.0.0.0.0.0.73.0",
    ("1-0:5.8.0", "24h"): "0.0.4.1.15.1.12.0.0.0.0.0.0.0.0.0.73.0",
    ("1-0:5.8.1", "24h"): "0.0.4.1.15.1.12.0.0.0.0.1.0.0.0.0.73.0",
    ("1-0:5.8.2", "24h"): "0.0.4.1.15.1.12.0.0.0.0.2.0.0.0.0.73.0",
    ("1-0:5.8.3", "24h"): "0.0.4.1.15.1.12.0.0.0.0.3.0.0.0.0.73.0",
    ("1-0:5.8.4", "24h"): "0.0.4.1.15.1.12.0.0.0.0.4.0.0.0.0.73.0",
    
    # Énergie réactive R2 (Rc+) - 6.8.x
    ("1-0:6.8.0", "15min"): "0.0.2.1.16.1.12.0.0.0.0.0.0.0.0.0.73.0",
    ("1-0:6.8.0", "24h"): "0.0.4.1.16.1.12.0.0.0.0.0.0.0.0.0.73.0",
    ("1-0:6.8.1", "24h"): "0.0.4.1.16.1.12.0.0.0.0.1.0.0.0.0.73.0",
    ("1-0:6.8.2", "24h"): "0.0.4.1.16.1.12.0.0.0.0.2.0.0.0.0.73.0",
    ("1-0:6.8.3", "24h"): "0.0.4.1.16.1.12.0.0.0.0.3.0.0.0.0.73.0",
    ("1-0:6.8.4", "24h"): "0.0.4.1.16.1.12.0.0.0.0.4.0.0.0.0.73.0",
    
    # Énergie réactive R3 (Ri-) - 7.8.x
    ("1-0:7.8.0", "15min"): "0.0.2.1.17.1.12.0.0.0.0.0.0.0.0.0.73.0",
    ("1-0:7.8.0", "24h"): "0.0.4.1.17.1.12.0.0.0.0.0.0.0.0.0.73.0",
    ("1-0:7.8.1", "24h"): "0.0.4.1.17.1.12.0.0.0.0.1.0.0.0.0.73.0",
    ("1-0:7.8.2", "24h"): "0.0.4.1.17.1.12.0.0.0.0.2.0.0.0.0.73.0",
    ("1-0:7.8.3", "24h"): "0.0.4.1.17.1.12.0.0.0.0.3.0.0.0.0.73.0",
    ("1-0:7.8.4", "24h"): "0.0.4.1.17.1.12.0.0.0.0.4.0.0.0.0.73.0",
    
    # Énergie réactive R4 (Rc-) - 8.8.x
    ("1-0:8.8.0", "15min"): "0.0.2.1.18.1.12.0.0.0.0.0.0.0.0.0.73.0",
    ("1-0:8.8.0", "24h"): "0.0.4.1.18.1.12.0.0.0.0.0.0.0.0.0.73.0",
    ("1-0:8.8.1", "24h"): "0.0.4.1.18.1.12.0.0.0.0.1.0.0.0.0.73.0",
    ("1-0:8.8.2", "24h"): "0.0.4.1.18.1.12.0.0.0.0.2.0.0.0.0.73.0",
    ("1-0:8.8.3", "24h"): "0.0.4.1.18.1.12.0.0.0.0.3.0.0.0.0.73.0",
    ("1-0:8.8.4", "24h"): "0.0.4.1.18.1.12.0.0.0.0.4.0.0.0.0.73.0",
}

# Descriptions OBIS pour affichage utilisateur
OBIS_DESCRIPTIONS = {
    "1-0:1.8.0": "A+ (Énergie active import)",
    "1-0:1.8.1": "A+T1",
    "1-0:1.8.2": "A+T2",
    "1-0:1.8.3": "A+T3",
    "1-0:1.8.4": "A+T4",
    "1-0:1.6.0": "Pmax(A+)T0 (Monthly power)",
    "1-0:2.8.0": "A- (Énergie active export)",
    "1-0:2.8.1": "A-T1",
    "1-0:2.8.2": "A-T2",
    "1-0:2.8.3": "A-T3",
    "1-0:2.8.4": "A-T4",
    "1-0:2.6.0": "Pmax(A-)T0 (Monthly power)",
    "1-0:5.8.0": "Ri+ (Énergie Réactive R1)",
    "1-0:5.8.1": "Ri+T1",
    "1-0:5.8.2": "Ri+T2",
    "1-0:5.8.3": "Ri+T3",
    "1-0:5.8.4": "Ri+T4",
    "1-0:6.8.0": "Rc+ (Énergie Réactive R2)",
    "1-0:6.8.1": "Rc+T1",
    "1-0:6.8.2": "Rc+T2",
    "1-0:6.8.3": "Rc+T3",
    "1-0:6.8.4": "Rc+T4",
    "1-0:7.8.0": "Ri- (Énergie Réactive R3)",
    "1-0:7.8.1": "Ri-T1",
    "1-0:7.8.2": "Ri-T2",
    "1-0:7.8.3": "Ri-T3",
    "1-0:7.8.4": "Ri-T4",
    "1-0:8.8.0": "Rc- (Énergie Réactive R4)",
    "1-0:8.8.1": "Rc-T1",
    "1-0:8.8.2": "Rc-T2",
    "1-0:8.8.3": "Rc-T3",
    "1-0:8.8.4": "Rc-T4",
}

# Mapping OBIS hex (pour XML) → OBIS lisible
OBIS_HEX_TO_READABLE = {
    # Clock
    "0000010000FF": "0-0:1.0.0",
    
    # Profile Status
    "0000600A01FF": "0-0:96.10.1",
    "0000600A02FF": "0-0:96.10.2",
    
    # Énergie active import (+A)
    "0100010800FF": "1-0:1.8.0",
    "0100010801FF": "1-0:1.8.1",
    "0100010802FF": "1-0:1.8.2",
    "0100010803FF": "1-0:1.8.3",
    "0100010804FF": "1-0:1.8.4",
    
    # Énergie active export (-A)
    "0100020800FF": "1-0:2.8.0",
    "0100020801FF": "1-0:2.8.1",
    "0100020802FF": "1-0:2.8.2",
    "0100020803FF": "1-0:2.8.3",
    "0100020804FF": "1-0:2.8.4",
    
    # Énergie réactive R1 (Ri+)
    "0100050800FF": "1-0:5.8.0",
    "0100050801FF": "1-0:5.8.1",
    "0100050802FF": "1-0:5.8.2",
    "0100050803FF": "1-0:5.8.3",
    "0100050804FF": "1-0:5.8.4",
    
    # Énergie réactive R2 (Rc+)
    "0100060800FF": "1-0:6.8.0",
    "0100060801FF": "1-0:6.8.1",
    "0100060802FF": "1-0:6.8.2",
    "0100060803FF": "1-0:6.8.3",
    "0100060804FF": "1-0:6.8.4",
    
    # Énergie réactive R3 (Ri-)
    "0100070800FF": "1-0:7.8.0",
    "0100070801FF": "1-0:7.8.1",
    "0100070802FF": "1-0:7.8.2",
    "0100070803FF": "1-0:7.8.3",
    "0100070804FF": "1-0:7.8.4",
    
    # Énergie réactive R4 (Rc-)
    "0100080800FF": "1-0:8.8.0",
    "0100080801FF": "1-0:8.8.1",
    "0100080802FF": "1-0:8.8.2",
    "0100080803FF": "1-0:8.8.3",
    "0100080804FF": "1-0:8.8.4",
}


# ============================================================================
# SECTION 2 : MODÈLES DE DONNÉES
# ============================================================================

@dataclass
class ParsedMeterData:
    """Données extraites d'un fichier de relevés de compteur."""
    meter_id: str
    load_profile: str
    interval: str  # "15min" ou "24h"
    channels: Dict[str, Dict[str, Any]]  # {obis: {"unit": str, "readings": [(ts, val, dst)]}}
    source_file: str
    warnings: List[str] = field(default_factory=list)
    needs_user_input: List[str] = field(default_factory=list)
    from_xml: bool = False  # Flag pour identifier les données issues de XML
    timestamps_utc: bool = False  # True si les horodatages sont déjà en UTC (ex: payload XML)


@dataclass
class ParseResult:
    """Résultat du parsing avec métadonnées de confiance."""
    data: Optional[ParsedMeterData]
    confidence: float  # 0.0 à 1.0
    strategy_used: str
    needs_user_input: List[str] = field(default_factory=list)


# ============================================================================
# SECTION 2B : STRUCTURES POUR PARSING XML FLEXIBLE
# ============================================================================

class XMLParseException(Exception):
    """Exception levée quand un fichier XML ne peut pas être parsé.

    Fournit des diagnostics détaillés pour aider l'utilisateur à comprendre
    pourquoi le parsing a échoué.
    """
    pass


@dataclass
class UnitInfo:
    """Informations sur une unité extraite du XML."""
    obis_code: str  # Format lisible (ex: "1-0:1.8.0")
    scaler: int  # Scaler DLMS (-3 = kilo, 0 = base, etc.)
    quantity: str  # Type de quantité (ActiveEnergy, ReactiveEnergy, etc.)
    resolved_unit: str  # Unité résolue (Wh, kWh, varh, etc.)


@dataclass
class ProfileObject:
    """Informations sur un profil découvert dans le XML."""
    object_name: str  # Nom de l'objet (ex: "DD.Profile_Load1")
    logical_name: str  # Code OBIS en hexadécimal
    profile_index: Optional[int]  # 1, 2, etc. (None si non détecté)
    class_id: int  # ClassID DLMS
    has_buffer: bool  # Possède attribut buffer
    has_capture_objects: bool  # Possède attribut capture_objects
    capture_period: Optional[int]  # Période de capture en secondes (si trouvée)
    buffer_patterns: List[str] = field(default_factory=list)  # Patterns de chemins découverts


@dataclass
class StructureMap:
    """Carte complète de la structure XML découverte."""
    meter_id: str  # DDID
    ddsubset: str  # ProfileBuffer, BillingValues, etc.
    profiles: List[ProfileObject]  # Profils découverts
    units: Dict[str, UnitInfo]  # OBIS → UnitInfo
    capture_objects: Dict[str, Dict[int, str]]  # profile_name → {idx: OBIS}
    buffer_patterns: Dict[str, List[re.Pattern]]  # profile_name → patterns regex
    timestamp_field_types: Dict[str, str]  # profile_name → field type
    namespace: Optional[str] = None  # Namespace XML si présent


@dataclass
class ReadingRow:
    """Une ligne de mesures avec timestamp."""
    timestamp: datetime
    dst_value: int  # Valeur DST (0=STD, 8=DST)
    values: Dict[str, float]  # OBIS code → valeur


# ============================================================================
# SECTION 3 : UTILITAIRES
# ============================================================================

def parse_timestamp(ts_string: Any) -> Optional[datetime]:
    """Parse différents formats de timestamp."""
    if ts_string is None:
        return None
    
    if isinstance(ts_string, datetime):
        return ts_string
    
    ts_clean = str(ts_string).strip()
    if not ts_clean:
        return None
    
    formats = [
        "%d/%m/%Y %H:%M:%S",
        "%d.%m.%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d.%m.%Y %H:%M",
        "%Y-%m-%d %H:%M",
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(ts_clean, fmt)
        except ValueError:
            continue
    
    return None


def parse_dst_value(dst_string: Any) -> int:
    """Extrait la valeur DST depuis une chaîne."""
    if dst_string is None:
        return 0
    
    match = re.search(r"(\d+)", str(dst_string).strip())
    return int(match.group(1)) if match else 0


def resolve_offset(dst_value: int, force_utc: bool = False) -> str:
    """Détermine l'offset à appliquer (+00:00 en UTC forcé, sinon offset DST local)."""
    if force_utc:
        return "+00:00"
    return DST_OFFSETS.get(dst_value, "+01:00")


def format_timestamp_iso(dt: datetime, dst_value: int, force_utc: bool = False) -> str:
    """Formate un timestamp au format ISO 8601 en appliquant ou non l'offset local."""
    offset = resolve_offset(dst_value, force_utc)
    return dt.strftime(f"%Y-%m-%dT%H:%M:%S.0000000{offset}")


def detect_interval(timestamps, fallback_interval: Optional[str] = None):
    """Détecte l'intervalle (15min ou 24h) depuis une liste de timestamps."""
    ts = sorted(t for t in timestamps if t is not None)
    
    if len(ts) < 2:
        if fallback_interval in ("15min", "24h"):
            return fallback_interval
        return "15min"

    diffs = [
        (ts[i+1] - ts[i]).total_seconds()
        for i in range(len(ts)-1)
        if (ts[i+1] - ts[i]).total_seconds() > 0
    ]

    if not diffs:
        if fallback_interval in ("15min", "24h"):
            return fallback_interval
        return "15min"

    avg = sum(diffs) / len(diffs)
    return "24h" if avg >= 82800 else "15min"


def detect_delimiter(content: str) -> str:
    """Détecte le délimiteur CSV le plus probable."""
    lines = content.split('\n')[:10]
    sample = '\n'.join(lines)
    
    best_delim = ";"
    best_score = 0
    
    for delim in CSV_DELIMITERS:
        try:
            reader = csv.reader(StringIO(sample), delimiter=delim)
            rows = list(reader)
            
            if len(rows) < 2:
                continue
            
            col_counts = [len(row) for row in rows if row]
            if not col_counts:
                continue
            
            avg = sum(col_counts) / len(col_counts)
            variance = sum((c - avg) ** 2 for c in col_counts) / len(col_counts)
            
            score = avg / (1 + variance)
            
            if score > best_score and avg > 1:
                best_score = score
                best_delim = delim
                
        except Exception:
            continue
    
    return best_delim


def read_file_content(file_bytes: bytes) -> Tuple[str, str]:
    """Lit un fichier avec détection d'encodage."""
    detected = chardet.detect(file_bytes)
    encoding = detected.get('encoding', 'utf-8') or 'utf-8'
    
    try:
        content = file_bytes.decode(encoding)
    except (UnicodeDecodeError, LookupError):
        content = file_bytes.decode('utf-8', errors='replace')
        encoding = 'utf-8-fallback'
    
    return content, encoding


def extract_obis_code(header: str) -> str:
    """Extrait le code OBIS d'un header de colonne."""
    match = re.search(r'(\d-\d:\d+\.\d+\.\d+)', header)
    return match.group(1) if match else ""


def extract_unit(header: str) -> str:
    """Extrait l'unité d'un header de colonne."""
    match = re.search(r'\[(kWh|kvarh|kVAh)\]|\((kWh|kvarh|kVAh)\)', header, re.IGNORECASE)
    if match:
        return match.group(1) or match.group(2)
    return ""


def extract_load_profile(line: str) -> Tuple[str, str, str]:
    """Extrait le profil de charge depuis la ligne 2 du CSV."""
    obis_match = re.search(r'(\d-\d:\d+\.\d+\.\d+)', line)
    obis = obis_match.group(1) if obis_match else ""
    
    profile_match = re.search(r'Profil de charge\s*(\d+)', line, re.IGNORECASE)
    
    if obis in LOAD_PROFILES:
        name, interval = LOAD_PROFILES[obis]
        return obis, name, interval
    elif profile_match:
        num = profile_match.group(1)
        name = f"Profil de charge {num}"
        interval = "15min" if num == "1" else "24h"
        return obis, name, interval
    
    return obis, "", "15min"


def get_file_extension(filename: str) -> str:
    """Retourne l'extension du fichier en minuscules."""
    return Path(filename).suffix.lower()


def get_obis_display_name(obis_code: str, unit: str = "") -> str:
    """Retourne un nom d'affichage pour un code OBIS."""
    parts = obis_code.split(":")
    if len(parts) == 2:
        short = parts[1].split("*")[0]
    else:
        short = obis_code
    
    if unit:
        return f"{short} ({unit})"
    return short


def obis_hex_to_readable(hex_obis: str) -> str:
    """Convertit un code OBIS hexadécimal en format lisible."""
    if hex_obis in OBIS_HEX_TO_READABLE:
        return OBIS_HEX_TO_READABLE[hex_obis]
    
    if len(hex_obis) != 12:
        return hex_obis
    
    try:
        a = int(hex_obis[0:2], 16)
        b = int(hex_obis[2:4], 16)
        c = int(hex_obis[4:6], 16)
        d = int(hex_obis[6:8], 16)
        e = int(hex_obis[8:10], 16)
        f = int(hex_obis[10:12], 16)
        
        if f == 0xFF:
            return f"{a}-{b}:{c}.{d}.{e}"
        else:
            return f"{a}-{b}:{c}.{d}.{e}*{f}"
            
    except ValueError:
        return hex_obis


def decode_dlms_timestamp(hex_string: str, force_utc: bool = False) -> Tuple[Optional[datetime], int]:
    """
    Décode un timestamp DLMS (OctetString 12 octets hex) en datetime.
    
    Retourne: (datetime, dst_value)
    """
    if not hex_string or len(hex_string) != 24:
        return None, 0
    
    try:
        year = int(hex_string[0:4], 16)
        month = int(hex_string[4:6], 16)
        day = int(hex_string[6:8], 16)
        hour = int(hex_string[10:12], 16)
        minute = int(hex_string[12:14], 16)
        second = int(hex_string[14:16], 16)
        timezone_offset = int(hex_string[18:22], 16)
        status = int(hex_string[22:24], 16)
        
        if month == 0xFF or day == 0xFF:
            return None, 0
        
        dt = datetime(year, month, day, hour, minute, second)
        
        dst_value = 8 if (status & 0x08) else 0
        
        if not force_utc and timezone_offset != 0x8000:
            if timezone_offset >= 0x8000:
                offset_minutes = timezone_offset - 0x10000
            else:
                offset_minutes = timezone_offset
            if offset_minutes == 120:
                dst_value = 8
            elif offset_minutes == 60:
                dst_value = 0
        
        return dt, dst_value
        
    except (ValueError, OverflowError):
        return None, 0


def resolve_unit_from_scaler(quantity: str, scaler: int) -> str:
    """Déduit l'unité réelle depuis Quantity/Scaler DLMS."""
    base_units = {
        "ActiveEnergy": "Wh",
        "ReactiveEnergy": "varh",
        "ApparentEnergy": "VAh"
    }

    base = base_units.get(quantity, "")
    if not base:
        return ""

    # En DLMS, Scaler = -3 → k
    prefix_map = {
        -3: "k",
        -6: "M",
        -9: "G"
    }

    if scaler == 0:
        return base

    if scaler in prefix_map:
        return f"{prefix_map[scaler]}{base}"

    return f"{base}*10^{scaler}"


# ============================================================================
# SECTION 3B : HANDLERS POUR PARSING XML FLEXIBLE
# ============================================================================

class OBISCodeHandler:
    """
    Gestionnaire de codes OBIS avec capacités de conversion générique.
    Utilise les fonctions existantes (obis_hex_to_readable, resolve_unit_from_scaler)
    et ajoute des méthodes d'inférence intelligente.
    """

    @staticmethod
    def hex_to_readable(hex_obis: str) -> str:
        """
        Convertit un code OBIS hex en format lisible.
        Wrapper autour de la fonction existante.
        """
        return obis_hex_to_readable(hex_obis)

    @staticmethod
    def infer_unit_from_obis(obis_code: str, scaler: Optional[int] = None,
                             quantity: Optional[str] = None) -> str:
        """
        Infère l'unité d'un code OBIS de manière intelligente.

        Args:
            obis_code: Code OBIS format lisible (ex: "1-0:1.8.0")
            scaler: Scaler DLMS si disponible
            quantity: Quantity DLMS si disponible (plus fiable)

        Returns:
            Unité déduite (Wh, kWh, varh, etc.)
        """
        # Si on a quantity + scaler, utiliser la fonction existante
        if quantity and scaler is not None:
            return resolve_unit_from_scaler(quantity, scaler)

        # Sinon, inférer depuis la structure OBIS
        # Format: A-B:C.D.E ou A-B:C.D.E*F
        parts = obis_code.replace('-', ':').split(':')
        if len(parts) < 2:
            return ""

        # Extraire le byte C (type de mesure)
        c_parts = parts[1].split('.')
        if len(c_parts) < 2:
            return ""

        try:
            c = int(c_parts[0])  # 1=actif, 2=actif export, 5-8=réactif
            d = int(c_parts[1])  # 6=max demand, 7=instantané, 8=énergie (time integral)

            # Déterminer le type d'énergie
            if c in [1, 2]:  # Active
                base_unit = "Wh"
            elif c in [5, 6, 7, 8]:  # Reactive
                base_unit = "varh"
            elif c in [9]:  # Apparent
                base_unit = "VAh"
            else:
                base_unit = "Wh"  # Default

            # Appliquer le scaler si disponible
            if scaler is not None:
                if scaler == 0:
                    return base_unit
                elif scaler == -3:
                    return f"k{base_unit}"
                elif scaler == -6:
                    return f"M{base_unit}"
                else:
                    return f"{base_unit}*10^{scaler}"

            return base_unit

        except (ValueError, IndexError):
            return "Wh"  # Safe default

    @staticmethod
    def is_energy_obis(obis_code: str) -> bool:
        """
        Vérifie si un code OBIS représente une mesure d'énergie
        (vs status, timestamp, configuration).

        Args:
            obis_code: Code OBIS format lisible (ex: "1-0:1.8.0")

        Returns:
            True si c'est une mesure d'énergie
        """
        # Filtrer les codes non-énergie connus
        if obis_code.startswith("0-0:1."):  # Clock/date
            return False
        if obis_code.startswith("0-0:96."):  # Status/config
            return False

        # Extraire le byte C
        parts = obis_code.replace('-', ':').split(':')
        if len(parts) < 2:
            return False

        c_parts = parts[1].split('.')
        if len(c_parts) < 3:
            return False

        try:
            c = int(c_parts[0])  # Type de mesure
            d = int(c_parts[1])  # Méthode

            # C in 1-9 = mesures électriques
            # D = 8 (time integral = énergie) ou 6 (max demand)
            if c in range(1, 10) and d in [6, 8]:
                return True

            return False

        except (ValueError, IndexError):
            return False

    @staticmethod
    def get_display_name(obis_code: str) -> str:
        """
        Génère un nom d'affichage lisible pour un code OBIS.
        Tente d'abord le lookup, sinon génère depuis la structure.

        Args:
            obis_code: Code OBIS format lisible

        Returns:
            Nom d'affichage (ex: "Active Energy Import (Total)")
        """
        # Tenter le lookup existant
        display_name = get_obis_display_name(obis_code)
        if display_name != obis_code:  # Si trouvé dans OBIS_DESCRIPTIONS
            return display_name

        # Générer depuis la structure
        parts = obis_code.replace('-', ':').split(':')
        if len(parts) < 2:
            return obis_code

        c_parts = parts[1].split('.')
        if len(c_parts) < 3:
            return obis_code

        try:
            c = int(c_parts[0])
            e = int(c_parts[2])

            # Déterminer le type
            energy_type = {
                1: "Active Energy Import",
                2: "Active Energy Export",
                5: "Reactive Energy Q1 (Ri+)",
                6: "Reactive Energy Q2 (Rc+)",
                7: "Reactive Energy Q3 (Ri-)",
                8: "Reactive Energy Q4 (Rc-)",
                9: "Apparent Energy"
            }.get(c, f"Energy Type {c}")

            # Déterminer le tarif/registre
            if e == 0:
                tariff = "(Total)"
            elif e in [1, 2, 3, 4]:
                tariff = f"(Tariff {e})"
            else:
                tariff = f"(Register {e})"

            return f"{energy_type} {tariff}"

        except (ValueError, IndexError):
            return obis_code


class ProfileDetector:
    """Détecteur de métadonnées de profil sans valeurs hardcodées."""

    def __init__(self, profile_obj: ProfileObject):
        self.profile = profile_obj

    def detect_interval(self, timestamps: List[datetime]) -> str:
        """Calcule l'intervalle depuis les différences de timestamps."""
        if not timestamps or len(timestamps) < 2:
            # Fallback sur capture_period si disponible
            if self.profile.capture_period:
                return self._seconds_to_interval(self.profile.capture_period)
            return "15min"  # Default

        # Calculer les différences
        diffs = []
        sorted_ts = sorted(timestamps)
        for i in range(1, min(len(sorted_ts), 100)):  # Limiter à 100 échantillons
            diff = (sorted_ts[i] - sorted_ts[i-1]).total_seconds()
            if diff > 0:  # Ignorer les doublons
                diffs.append(diff)

        if not diffs:
            return "15min"

        # Trouver le mode (valeur la plus fréquente)
        from collections import Counter
        counter = Counter(diffs)
        most_common_diff = counter.most_common(1)[0][0]

        return self._seconds_to_interval(int(most_common_diff))

    def _seconds_to_interval(self, seconds: int) -> str:
        """Convertit secondes en format lisible."""
        if seconds == 900:
            return "15min"
        elif seconds == 1800:
            return "30min"
        elif seconds == 3600:
            return "1h"
        elif seconds == 21600:
            return "6h"
        elif seconds == 86400:
            return "24h"
        elif seconds < 3600:
            return f"{seconds//60}min"
        elif seconds < 86400:
            return f"{seconds//3600}h"
        else:
            return f"{seconds//86400}d"

    def detect_profile_name(self, interval: str) -> str:
        """Génère un nom de profil significatif."""
        # Extraire numéro depuis object_name si possible
        if self.profile.profile_index:
            return f"Profil de charge {self.profile.profile_index}"

        # Fallback sur interval
        return f"Profil de charge ({interval})"


class XMLStructureExplorer:
    """Explorateur de structure XML DLMS/COSEM."""

    def __init__(self, root: ET.Element):
        self.root = root
        self.namespace = self._detect_namespace()

    def _detect_namespace(self) -> Optional[str]:
        """Détecte le namespace XML."""
        if '}' in self.root.tag:
            return self.root.tag.split('}')[0][1:]
        return None

    def discover_structure(self, ddid: str, ddsubset: str) -> StructureMap:
        """Découvre la structure complète du XML."""
        profiles = self.find_profile_objects()
        units = self.extract_units_from_registers()

        # Pour chaque profil, découvrir capture_objects et buffer_patterns
        capture_objects_map = {}
        buffer_patterns_map = {}

        for profile in profiles:
            capture_objs = self.find_capture_objects_dynamic(profile.object_name)
            capture_objects_map[profile.object_name] = capture_objs

            buffer_pats = self.find_buffer_data_paths(profile.object_name)
            buffer_patterns_map[profile.object_name] = buffer_pats

        return StructureMap(
            meter_id=ddid,
            ddsubset=ddsubset,
            profiles=profiles,
            units=units,
            capture_objects=capture_objects_map,
            buffer_patterns=buffer_patterns_map,
            timestamp_field_types={},
            namespace=self.namespace
        )

    def find_profile_objects(self) -> List[ProfileObject]:
        """Trouve les profils Load1 et Load2."""
        profiles = []

        for obj in self.root.iter():
            tag_name = obj.tag.split('}')[-1] if '}' in obj.tag else obj.tag
            if tag_name != "Objects":
                continue

            obj_name = obj.get("ObjectName", "")
            if not obj_name:
                continue

            # Chercher Load1 ou Load2 (variations: Load1, Load01, Load_1, etc.)
            profile_num = None
            if re.search(r'(Profile|Load)[\s_]*0?1\b', obj_name, re.IGNORECASE):
                profile_num = 1
            elif re.search(r'(Profile|Load)[\s_]*0?2\b', obj_name, re.IGNORECASE):
                profile_num = 2

            if profile_num is None:
                continue

            # Extraire métadonnées
            logical_name = obj.get("ObjectLogicalName", "")
            class_id = int(obj.get("ClassID", "0"))

            # Vérifier présence buffer et capture_objects
            has_buffer = False
            has_capture_objects = False
            capture_period = None

            for attr in obj.iter():
                attr_name = attr.get("AttributeName", "")
                if "buffer" in attr_name.lower():
                    has_buffer = True
                if "capture_objects" in attr_name.lower():
                    has_capture_objects = True
                if "capture_period" in attr_name.lower():
                    field_val = attr.get("FieldValue", "")
                    try:
                        capture_period = int(field_val)
                    except (ValueError, TypeError):
                        pass

            if has_buffer and has_capture_objects:
                profiles.append(ProfileObject(
                    object_name=obj_name,
                    logical_name=logical_name,
                    profile_index=profile_num,
                    class_id=class_id,
                    has_buffer=has_buffer,
                    has_capture_objects=has_capture_objects,
                    capture_period=capture_period
                ))

        return profiles

    def find_capture_objects_dynamic(self, profile_name: str) -> Dict[int, str]:
        """Trouve les capture_objects sans assumer la structure."""
        capture_objects = {}

        # Pattern flexible pour capturer tous les chemins possibles
        # Ex: DD.Profile_Load1.capture_objects.0.2.logical_name
        pattern = re.compile(rf".*{re.escape(profile_name)}.*capture_objects.*\.(\d+)\.logical_name$")

        for field in self.root.iter():
            field_name = field.get("FieldName", "")
            match = pattern.search(field_name)
            if match:
                idx = int(match.group(1))
                obis_hex = field.get("FieldValue", "")
                if obis_hex:
                    obis_readable = OBISCodeHandler.hex_to_readable(obis_hex)
                    capture_objects[idx] = obis_readable

        return capture_objects

    def find_buffer_data_paths(self, profile_name: str) -> List[re.Pattern]:
        """Découvre dynamiquement les patterns de buffer."""
        patterns_found = set()

        # Chercher tous les champs buffer pour ce profil
        for field in self.root.iter():
            field_name = field.get("FieldName", "")

            if profile_name not in field_name or "buffer" not in field_name:
                continue

            # Essayer de matcher différents patterns
            # Pattern 1: buffer.Selector1.Response.X.Y
            match1 = re.search(rf"{re.escape(profile_name)}\.buffer\.(Selector\d+\.Response)\.(\d+)\.(\d+)$", field_name)
            if match1:
                patterns_found.add(f"Selector{match1.group(1).split('Selector')[1].split('.')[0]}.Response")

            # Pattern 2: buffer.0.X.Y
            match2 = re.search(rf"{re.escape(profile_name)}\.buffer\.(\d+)\.(\d+)\.(\d+)$", field_name)
            if match2:
                patterns_found.add(f"{match2.group(1)}")

        # Convertir en regex patterns
        regex_patterns = []
        for pat_str in patterns_found:
            if "Selector" in pat_str:
                regex_patterns.append(
                    re.compile(rf"{re.escape(profile_name)}\.buffer\.{re.escape(pat_str)}\.(\d+)\.(\d+)$")
                )
            else:
                regex_patterns.append(
                    re.compile(rf"{re.escape(profile_name)}\.buffer\.{re.escape(pat_str)}\.(\d+)\.(\d+)$")
                )

        return regex_patterns

    def extract_units_from_registers(self) -> Dict[str, UnitInfo]:
        """Extrait unités de tous les registres."""
        units = {}

        for obj in self.root.iter():
            tag_name = obj.tag.split('}')[-1] if '}' in obj.tag else obj.tag
            if tag_name != "Objects":
                continue

            class_id = obj.get("ClassID", "")
            if class_id != "3":  # Pas un register
                continue

            obj_logical_name = obj.get("ObjectLogicalName", "")
            if not obj_logical_name:
                continue

            # Chercher scaler_unit dans les attributes
            scaler = None
            quantity = None

            for attr in obj.iter():
                attr_name = attr.get("AttributeName", "")
                if "scaler_unit" in attr_name.lower() or "UnitScale" in attr_name:
                    for field in attr.iter():
                        field_name = field.get("FieldName", "")
                        if "Scaler" in field_name:
                            try:
                                scaler = int(field.get("FieldValue", "0"))
                            except (ValueError, TypeError):
                                pass
                        elif "Quantity" in field_name:
                            quantity = field.get("FieldValue", "")

            if scaler is not None and quantity:
                obis_readable = OBISCodeHandler.hex_to_readable(obj_logical_name)
                resolved_unit = OBISCodeHandler.infer_unit_from_obis(obis_readable, scaler, quantity)

                units[obis_readable] = UnitInfo(
                    obis_code=obis_readable,
                    scaler=scaler,
                    quantity=quantity,
                    resolved_unit=resolved_unit
                )

        return units


class DataPathResolver:
    """Résolveur de chemins de données dans le XML."""

    def __init__(self, root: ET.Element, structure_map: StructureMap):
        self.root = root
        self.structure = structure_map

    def extract_buffer_data(self, profile_name: str) -> Dict[int, Dict[int, str]]:
        """Extrait les données du buffer."""
        buffer_data = {}
        patterns = self.structure.buffer_patterns.get(profile_name, [])

        if not patterns:
            return buffer_data

        for field in self.root.iter():
            field_name = field.get("FieldName", "")

            for pattern in patterns:
                match = pattern.match(field_name)
                if match:
                    row = int(match.group(1))
                    col = int(match.group(2))
                    value = field.get("FieldValue", "")

                    if row not in buffer_data:
                        buffer_data[row] = {}
                    buffer_data[row][col] = value
                    break

        return buffer_data

    def match_timestamps_to_values(self, buffer_data: Dict[int, Dict[int, str]],
                                   capture_objects: Dict[int, str]) -> List[ReadingRow]:
        """Associe timestamps aux valeurs."""
        readings = []

        for row_idx in sorted(buffer_data.keys()):
            row_data = buffer_data[row_idx]

            # Column 0 = timestamp
            if 0 not in row_data:
                continue

            timestamp_hex = row_data[0]
            dt, dst_val = decode_dlms_timestamp(timestamp_hex)

            if dt is None:
                continue

            # Extraire valeurs pour chaque OBIS
            values = {}
            for col_idx, obis_code in capture_objects.items():
                if col_idx < 2:  # Skip timestamp et DST
                    continue

                if col_idx in row_data:
                    try:
                        values[obis_code] = float(row_data[col_idx])
                    except (ValueError, TypeError):
                        pass

            if values:
                readings.append(ReadingRow(
                    timestamp=dt,
                    dst_value=dst_val,
                    values=values
                ))

        return readings


# ============================================================================
# SECTION 4 : PARSERS
# ============================================================================

def extract_meter_id_smart(lines: List[str], filename: str, delimiter: str = ";") -> Tuple[Optional[str], int, str]:
    """Recherche intelligente du numéro de compteur."""
    
    for idx, line in enumerate(lines[:10]):
        cleaned = line.strip().replace('\ufeff', '')
        if re.match(r'^\d{6,}$', cleaned):
            return cleaned, idx, "ligne_seule"
    
    patterns_label = [
        r'(?:compteur|meter|serial|numéro|n°|id)[:\s=]+(\d{6,})',
        r'(?:device|appareil|zähler)[:\s=]+(\d{6,})',
        r'^[^;,\t]*?(\d{6,})[^;,\t]*$',
    ]
    
    for idx, line in enumerate(lines[:15]):
        for pattern in patterns_label:
            match = re.search(pattern, line, re.IGNORECASE)
            if match:
                return match.group(1), idx, "pattern_label"
    
    keywords = ['compteur', 'meter', 'serial', 'numéro', 'device', 'id', 'zähler', 'appareil']
    
    for idx, line in enumerate(lines[:20]):
        cells = [c.strip().lower() for c in line.split(delimiter)]
        for i, cell in enumerate(cells):
            if any(kw in cell for kw in keywords):
                for j in range(max(0, i-1), min(len(cells), i+3)):
                    num_match = re.search(r'(\d{6,})', cells[j])
                    if num_match:
                        return num_match.group(1), idx, "tableau_cle_valeur"
    
    filename_clean = Path(filename).stem
    filename_patterns = [
        r'^(\d{6,})',
        r'[_\-](\d{6,})[_\-]',
        r'[_\-](\d{6,})$',
        r'(\d{8,})',
    ]
    
    for pattern in filename_patterns:
        match = re.search(pattern, filename_clean)
        if match:
            return match.group(1), -1, "nom_fichier"
    
    return None, -1, "non_trouvé"


def detect_file_structure(lines: List[str], delimiter: str) -> Dict[str, Any]:
    """Détecte automatiquement la structure du fichier."""
    structure = {
        "header_line": None,
        "data_start": None,
        "profile_line": None,
        "meter_id_line": None,
        "confidence": 0.0
    }
    
    header_candidates = []
    data_candidates = []
    
    for idx, line in enumerate(lines[:30]):
        if not line.strip():
            continue
            
        cells = [c.strip() for c in line.split(delimiter)]
        num_cells = len([c for c in cells if c])
        
        obis_count = sum(1 for c in cells if re.search(r'\d-\d:\d+\.\d+\.\d+', c))
        header_keywords = sum(1 for c in cells if re.search(
            r'horloge|timestamp|date|heure|time|état|status|valeur|value|energie|energy',
            c, re.IGNORECASE
        ))
        
        if obis_count >= 2 or (header_keywords >= 2 and num_cells >= 3):
            header_candidates.append((idx, obis_count + header_keywords))
        
        has_timestamp = any(parse_timestamp(c) for c in cells[:5])
        numeric_count = sum(1 for c in cells if re.match(r'^[\d,.\-]+$', c.strip()) and c.strip())
        
        if has_timestamp and numeric_count >= 2:
            data_candidates.append(idx)
        
        if re.search(r'1-0:99\.\d+\.0|profil de charge|load profile', line, re.IGNORECASE):
            structure["profile_line"] = idx
    
    if header_candidates:
        header_candidates.sort(key=lambda x: x[1], reverse=True)
        structure["header_line"] = header_candidates[0][0]
    
    if data_candidates:
        if structure["header_line"] is not None:
            valid_data = [d for d in data_candidates if d > structure["header_line"]]
            structure["data_start"] = valid_data[0] if valid_data else data_candidates[0]
        else:
            structure["data_start"] = data_candidates[0]
            if data_candidates[0] > 0:
                structure["header_line"] = data_candidates[0] - 1
    
    confidence = 0.0
    if structure["header_line"] is not None:
        confidence += 0.4
    if structure["data_start"] is not None:
        confidence += 0.4
    if structure["profile_line"] is not None:
        confidence += 0.1
    if structure["header_line"] is not None and structure["data_start"] is not None:
        if structure["data_start"] == structure["header_line"] + 1:
            confidence += 0.1
    
    structure["confidence"] = confidence
    return structure


def validate_parse_result(data: ParsedMeterData) -> Tuple[bool, List[str]]:
    """Valide qu'un résultat de parsing est utilisable."""
    problems = []
    
    if not data.meter_id:
        problems.append("meter_id")
    
    if not data.channels:
        problems.append("channels")
    else:
        total_readings = sum(len(ch["readings"]) for ch in data.channels.values())
        if total_readings == 0:
            problems.append("readings")
    
    is_valid = len(problems) == 0
    return is_valid, problems


def _build_column_map(headers: List[str]) -> Tuple[Dict[int, Dict], Dict[str, Dict]]:
    """Construit le mapping des colonnes depuis les headers."""
    col_map: Dict[int, Dict[str, str]] = {}
    channels: Dict[str, Dict[str, Any]] = {}
    
    for idx, header in enumerate(headers):
        obis = extract_obis_code(header)
        unit = extract_unit(header)
        
        if "0-0:1.0.0" in header or re.search(r'horloge|timestamp|date.?time', header, re.IGNORECASE):
            col_map[idx] = {"type": "timestamp", "obis": "", "unit": ""}
        elif "0-0:96.10.1" in header or re.search(r'état|status|dst', header, re.IGNORECASE):
            col_map[idx] = {"type": "dst", "obis": "", "unit": ""}
        elif obis:
            col_map[idx] = {"type": "data", "obis": obis, "unit": unit}
            channels[obis] = {"unit": unit, "readings": []}
    
    return col_map, channels


def _infer_columns_from_data(
    headers: List[str], 
    sample_lines: List[str], 
    delimiter: str
) -> Tuple[Dict[int, Dict], Dict[str, Dict]]:
    """Devine le type des colonnes en analysant les données."""
    col_map: Dict[int, Dict] = {}
    channels: Dict[str, Dict] = {}
    
    if not sample_lines:
        return col_map, channels
    
    num_cols = len(headers)
    
    for col_idx in range(num_cols):
        values = []
        for line in sample_lines:
            cells = line.split(delimiter)
            if col_idx < len(cells):
                values.append(cells[col_idx].strip())
        
        timestamp_count = sum(1 for v in values if parse_timestamp(v))
        numeric_count = sum(1 for v in values if re.match(r'^[\d,.\-]+$', v) and v)
        dst_count = sum(1 for v in values if re.match(r'^\d+\s*\(?(DST|STD)?\)?$', v, re.I))
        
        total = len(values)
        
        if total > 0:
            if timestamp_count > total * 0.7:
                col_map[col_idx] = {"type": "timestamp", "obis": "", "unit": ""}
            elif dst_count > total * 0.7:
                col_map[col_idx] = {"type": "dst", "obis": "", "unit": ""}
            elif numeric_count > total * 0.8:
                header_clean = re.sub(r'[^\w]', '_', headers[col_idx])[:20]
                pseudo_obis = f"unknown:{header_clean}"
                col_map[col_idx] = {"type": "data", "obis": pseudo_obis, "unit": ""}
                channels[pseudo_obis] = {"unit": "", "readings": []}
    
    return col_map, channels


def _parse_data_rows(
    data_lines: List[str],
    delimiter: str,
    col_map: Dict[int, Dict],
    channels: Dict[str, Dict],
    ts_idx: Optional[int],
    dst_idx: Optional[int]
) -> None:
    """Parse les lignes de données et remplit les channels."""
    for line in data_lines:
        line = line.strip()
        if not line:
            continue
        
        values = [v.strip() for v in line.split(delimiter)]
        
        timestamp = None
        if ts_idx is not None and ts_idx < len(values):
            timestamp = parse_timestamp(values[ts_idx])
        
        if timestamp is None:
            continue
        
        dst_val = 0
        if dst_idx is not None and dst_idx < len(values):
            dst_val = parse_dst_value(values[dst_idx])
        
        for idx, col_info in col_map.items():
            if col_info["type"] != "data":
                continue
            
            if idx >= len(values):
                continue
            
            val_str = values[idx].strip()
            if not val_str:
                continue
            
            try:
                value = float(val_str.replace(",", "."))
                obis = col_info["obis"]
                if obis in channels:
                    channels[obis]["readings"].append((timestamp, value, dst_val))
            except ValueError:
                pass


def parse_csv_standard(lines: List[str], delimiter: str, filename: str) -> ParseResult:
    """Parser CSV format standard (structure fixe)."""
    warnings: List[str] = []
    
    if len(lines) < 4:
        return ParseResult(None, 0.0, "standard", ["structure"])
    
    meter_id = lines[0].strip().replace('\ufeff', '')
    
    if not re.match(r'^\d+$', meter_id):
        return ParseResult(None, 0.0, "standard", ["meter_id"])
    
    obis_profile, profile_name, interval = extract_load_profile(lines[1])
    
    if not profile_name:
        warnings.append("Profil de charge non détecté, utilisation par défaut: 15min")
        profile_name = "Profil de charge 1"
        interval = "15min"
    
    headers = [h.strip() for h in lines[2].split(delimiter)]
    
    if not any(re.search(r'\d-\d:\d+\.\d+\.\d+', h) for h in headers):
        return ParseResult(None, 0.3, "standard", ["headers"])
    
    col_map, channels = _build_column_map(headers)
    ts_idx = next((i for i, c in col_map.items() if c["type"] == "timestamp"), None)
    dst_idx = next((i for i, c in col_map.items() if c["type"] == "dst"), None)
    
    if ts_idx is None:
        warnings.append("Colonne timestamp non trouvée")
    
    _parse_data_rows(lines[3:], delimiter, col_map, channels, ts_idx, dst_idx)
    
    all_timestamps = [r[0] for ch in channels.values() for r in ch["readings"]]
    detected_interval = detect_interval(all_timestamps, interval)
    
    if detected_interval != interval:
        warnings.append(f"Intervalle détecté ({detected_interval}) différent du profil ({interval})")
        interval = detected_interval
    
    data = ParsedMeterData(
        meter_id=meter_id,
        load_profile=profile_name,
        interval=interval,
        channels=channels,
        source_file=filename,
        warnings=warnings
    )
    
    is_valid, problems = validate_parse_result(data)
    confidence = 1.0 if is_valid else 0.5
    
    return ParseResult(data, confidence, "standard", problems if not is_valid else [])


def parse_csv_heuristic(lines: List[str], delimiter: str, filename: str) -> ParseResult:
    """Parser CSV par détection heuristique (fallback)."""
    warnings: List[str] = []
    needs_input: List[str] = []
    
    structure = detect_file_structure(lines, delimiter)
    
    if structure["header_line"] is None or structure["data_start"] is None:
        return ParseResult(None, 0.0, "heuristic", ["structure"])
    
    meter_id, meter_line, meter_strategy = extract_meter_id_smart(lines, filename, delimiter)
    
    if meter_id:
        warnings.append(f"Compteur détecté via: {meter_strategy}")
    else:
        needs_input.append("meter_id")
        meter_id = ""
    
    profile_name = ""
    interval = "15min"
    
    if structure["profile_line"] is not None:
        obis_profile, profile_name, interval = extract_load_profile(lines[structure["profile_line"]])
    
    if not profile_name:
        warnings.append("Profil non détecté, intervalle sera déduit des données")
    
    headers = [h.strip() for h in lines[structure["header_line"]].split(delimiter)]
    col_map, channels = _build_column_map(headers)
    
    if not any(c["type"] == "data" for c in col_map.values()):
        warnings.append("Headers sans codes OBIS, tentative de détection par contenu")
        col_map, channels = _infer_columns_from_data(
            headers, 
            lines[structure["data_start"]:structure["data_start"]+10],
            delimiter
        )
    
    ts_idx = next((i for i, c in col_map.items() if c["type"] == "timestamp"), None)
    dst_idx = next((i for i, c in col_map.items() if c["type"] == "dst"), None)
    
    if ts_idx is None:
        needs_input.append("timestamp_column")
    
    _parse_data_rows(lines[structure["data_start"]:], delimiter, col_map, channels, ts_idx, dst_idx)
    
    all_timestamps = [r[0] for ch in channels.values() for r in ch["readings"]]
    if all_timestamps:
        interval = detect_interval(all_timestamps, interval if profile_name else None)
    
    data = ParsedMeterData(
        meter_id=meter_id,
        load_profile=profile_name or f"Détecté: {interval}",
        interval=interval,
        channels=channels,
        source_file=filename,
        warnings=warnings
    )
    
    is_valid, problems = validate_parse_result(data)
    needs_input.extend(problems)
    
    confidence = structure["confidence"] * (0.8 if meter_id else 0.5)
    
    return ParseResult(data, confidence, "heuristic", list(set(needs_input)))


def parse_csv(file_bytes: bytes, filename: str, delimiter: str = ";", 
              user_config: Optional[Dict] = None) -> List[ParsedMeterData]:
    """Parse un fichier CSV avec stratégie adaptative."""
    content, encoding = read_file_content(file_bytes)
    lines = content.strip().split('\n')
    
    if not lines:
        return []
    
    if delimiter == ";":
        detected_delim = detect_delimiter(content)
    else:
        detected_delim = delimiter
    
    result = parse_csv_standard(lines, detected_delim, filename)
    
    if result.confidence >= 0.8 and not result.needs_user_input:
        return [result.data] if result.data else []
    
    fallback_result = parse_csv_heuristic(lines, detected_delim, filename)
    
    if fallback_result.confidence > result.confidence:
        result = fallback_result
    
    if user_config and result.data:
        if "meter_id" in user_config and user_config["meter_id"]:
            result.data.meter_id = user_config["meter_id"]
            if "meter_id" in result.needs_user_input:
                result.needs_user_input.remove("meter_id")
        
        if "interval" in user_config and user_config["interval"]:
            result.data.interval = user_config["interval"]
    
    if result.data:
        if result.needs_user_input:
            result.data.needs_user_input = result.needs_user_input
            result.data.warnings.append(
                f"⚠️ Configuration manuelle requise: {', '.join(result.needs_user_input)}"
            )
        return [result.data]
    
    return []


def parse_xlsx(file_bytes: bytes, filename: str) -> List[ParsedMeterData]:
    """Parse un fichier Excel de relevés de compteur."""
    warnings: List[str] = []
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        warnings.append(f"Erreur de lecture Excel: {str(e)}")
        return []
    
    meter_id = ""
    profile_name = ""
    interval = ""
    
    info_sheet = None
    for name in wb.sheetnames:
        name_lower = name.lower()
        if "info" in name_lower or "général" in name_lower or "general" in name_lower:
            info_sheet = wb[name]
            break
    
    if info_sheet is None and len(wb.sheetnames) > 0:
        info_sheet = wb[wb.sheetnames[0]]
    
    if info_sheet:
        for row in info_sheet.iter_rows(min_row=1, max_row=20, max_col=3):
            cell_a = str(row[0].value or "").lower()
            cell_b = str(row[1].value or "") if len(row) > 1 else ""
            
            if "numéro de série" in cell_a or "serial" in cell_a or "numéro" in cell_a:
                meter_id = cell_b.strip()
            elif "profil" in cell_a or "profile" in cell_a:
                profile_name = cell_b.strip()
                if "1" in profile_name:
                    interval = "15min"
                elif "2" in profile_name:
                    interval = "24h"
    
    if not meter_id:
        warnings.append("Numéro de série non trouvé dans l'onglet d'informations")
    
    data_sheet = None
    for name in wb.sheetnames:
        name_lower = name.lower()
        if "donnée" in name_lower or "data" in name_lower or "profile" in name_lower:
            data_sheet = wb[name]
            break
    
    if data_sheet is None and len(wb.sheetnames) > 1:
        data_sheet = wb[wb.sheetnames[1]]
    
    if data_sheet is None:
        warnings.append("Onglet de données non trouvé")
        return []
    
    headers = [str(cell.value or "") for cell in data_sheet[1]]
    
    col_map: Dict[int, Dict[str, str]] = {}
    channels: Dict[str, Dict[str, Any]] = {}
    
    for idx, header in enumerate(headers):
        obis = extract_obis_code(header)
        unit = extract_unit(header)
        
        header_lower = header.lower()
        
        if "0-0:1.0.0" in header or "horloge" in header_lower:
            col_map[idx] = {"type": "timestamp", "obis": "", "unit": ""}
        elif "0-0:96.10.1" in header or "état" in header_lower or "status" in header_lower:
            col_map[idx] = {"type": "dst", "obis": "", "unit": ""}
        elif obis:
            col_map[idx] = {"type": "data", "obis": obis, "unit": unit}
            channels[obis] = {"unit": unit, "readings": []}
    
    ts_idx = next((i for i, c in col_map.items() if c["type"] == "timestamp"), None)
    dst_idx = next((i for i, c in col_map.items() if c["type"] == "dst"), None)
    
    for row in data_sheet.iter_rows(min_row=2):
        timestamp = None
        if ts_idx is not None and ts_idx < len(row):
            ts_val = row[ts_idx].value
            if isinstance(ts_val, datetime):
                timestamp = ts_val
            elif ts_val:
                timestamp = parse_timestamp(str(ts_val))
        
        if timestamp is None:
            continue
        
        dst_val = 0
        if dst_idx is not None and dst_idx < len(row):
            dst_val = parse_dst_value(str(row[dst_idx].value or ""))
        
        for idx, col_info in col_map.items():
            if col_info["type"] != "data":
                continue
            
            if idx >= len(row):
                continue
            
            cell_val = row[idx].value
            if cell_val is None:
                continue
            
            try:
                if isinstance(cell_val, str):
                    value = float(cell_val.replace(",", "."))
                else:
                    value = float(cell_val)
                
                obis = col_info["obis"]
                channels[obis]["readings"].append((timestamp, value, dst_val))
            except (ValueError, TypeError):
                pass
    
    all_timestamps = [r[0] for ch in channels.values() for r in ch["readings"]]
    detected_interval = detect_interval(all_timestamps, interval)
    
    if detected_interval != interval:
        warnings.append(f"Intervalle détecté ({detected_interval}) différent du profil ({interval})")
        interval = detected_interval
    
    return [ParsedMeterData(
        meter_id=meter_id,
        load_profile=profile_name,
        interval=interval,
        channels=channels,
        source_file=filename,
        warnings=warnings
    )]


def parse_xml(file_bytes: bytes, filename: str) -> List[ParsedMeterData]:
    """
    Parse un fichier XML DLMS/COSEM (ProfileBuffer ou BillingValues).

    Points importants:
    - Le meter_id est extrait de <DDID> et contient déjà l'info complète
    - Pas de concaténation avec le préfixe mRID manuel
    - L'unité est déduite du scaler_unit DLMS (Wh, kWh, varh, etc.)
    - Détecte automatiquement le type de fichier (ProfileBuffer vs BillingValues)
    """
    warnings: List[str] = []

    try:
        content = file_bytes.decode('utf-8')
    except UnicodeDecodeError:
        try:
            content = file_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = file_bytes.decode('latin-1')

    try:
        root = ET.fromstring(content)
    except ET.ParseError as e:
        warnings.append(f"Erreur de parsing XML: {str(e)}")
        return []

    # === Détection du type de fichier ===
    # Vérifier l'attribut DDSubset dans l'élément DDs
    dd_subset = None

    # Chercher avec namespace
    ns = {"ns": "http://tempuri.org/DeviceDescriptionDataSet.xsd"}
    dds_elem = root.find(".//ns:DDs", ns)
    if dds_elem is not None:
        dd_subset = dds_elem.get("DDSubset", "")

    # Fallback sans namespace
    if dd_subset is None:
        for elem in root.iter():
            tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
            if tag_name == "DDs":
                dd_subset = elem.get("DDSubset", "")
                break

    # Router vers la bonne fonction de parsing avec fallback flexible
    try:
        if dd_subset == "BillingValues":
            result = parse_xml_billing_values(file_bytes, filename)
        else:
            # Par défaut, traiter comme ProfileBuffer
            result = parse_xml_profile_buffer(file_bytes, filename)

        # COUCHE 2: Fallback vers parser flexible si résultat vide
        if not result or (result and len(result[0].channels) == 0):
            # Parser existant n'a pas réussi, tenter découverte flexible
            result = parse_xml_flexible(file_bytes, filename, dd_subset or "ProfileBuffer")

        return result

    except XMLParseException:
        # Re-raise les erreurs de parsing flexible (fail completely)
        raise

    except Exception as e:
        # Autres erreurs: tenter fallback avant d'échouer
        try:
            return parse_xml_flexible(file_bytes, filename, dd_subset or "ProfileBuffer")
        except XMLParseException:
            raise
        except Exception as flex_error:
            # Les deux parsers ont échoué
            raise XMLParseException(
                f"Impossible de parser {filename}\n\n"
                f"Raison: Échec du parser standard ET du parser flexible\n"
                f"DDSubset: {dd_subset or 'ProfileBuffer'}\n"
                f"Erreur standard: {type(e).__name__}: {str(e)}\n"
                f"Erreur flexible: {type(flex_error).__name__}: {str(flex_error)}\n\n"
                f"Suggestion: Structure XML non supportée.\n"
                f"Contactez le support technique avec ce fichier."
            )


def parse_xml_profile_buffer(file_bytes: bytes, filename: str) -> List[ParsedMeterData]:
    """
    Parse un fichier XML DLMS/COSEM de profil de charge (ProfileBuffer).
    Utilisé pour les fichiers E360, LGZ avec des séries temporelles.
    """
    warnings: List[str] = []

    try:
        content = file_bytes.decode('utf-8')
    except UnicodeDecodeError:
        try:
            content = file_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = file_bytes.decode('latin-1')

    try:
        root = ET.fromstring(content)
    except ET.ParseError as e:
        warnings.append(f"Erreur de parsing XML: {str(e)}")
        return []

    # === 1. Extraire DDID (meter ID) ===
    ddid_elem = root.find(".//{http://tempuri.org/DeviceDescriptionDataSet.xsd}DDID")
    if ddid_elem is None:
        ddid_elem = root.find(".//DDID")
    
    if ddid_elem is None or not ddid_elem.text:
        warnings.append("DDID (meter ID) non trouvé dans le fichier XML")
        return []
    
    meter_id = ddid_elem.text.strip()
    
    # === 2. Détecter le profil de charge ===
    profile_name = ""
    interval = ""
    profile_num = ""
    
    for obj in root.iter():
        obj_name = obj.get("ObjectName", "")
        if "ProfileStatus_Load01" in obj_name:
            profile_name = "Profil de charge 1"
            interval = "15min"
            profile_num = "1"
            break
        elif "ProfileStatus_Load02" in obj_name:
            profile_name = "Profil de charge 2"
            interval = "24h"
            profile_num = "2"
            break
    
    if not profile_name:
        for field in root.iter():
            field_name = field.get("FieldName", "")
            if "capture_period" in field_name and field.get("FieldValue"):
                try:
                    period = int(field.get("FieldValue", "0"))
                    if period == 900:
                        profile_name = "Profil de charge 1"
                        interval = "15min"
                        profile_num = "1"
                    elif period == 86400:
                        profile_name = "Profil de charge 2"
                        interval = "24h"
                        profile_num = "2"
                except ValueError:
                    pass
                break
    
    if not profile_name:
        warnings.append("Type de profil non détecté, utilisation par défaut: 15min")
        profile_name = "Profil de charge"
        interval = "15min"
        profile_num = "1"

    # === 3. Extraire les unités (scaler_unit) des TariffEnergyRegister ===
    tariff_registers: Dict[str, Dict[str, Any]] = {}

    pattern_tariff_logical = re.compile(r"DD\.TariffEnergyRegister_(\d+)\.logical_name")
    pattern_tariff_scaler = re.compile(r"DD\.TariffEnergyRegister_(\d+)\.scaler_unit\.Scaler")
    pattern_tariff_quantity = re.compile(r"DD\.TariffEnergyRegister_(\d+)\.scaler_unit\.Quantity")

    for field in root.iter():
        field_name = field.get("FieldName", "")
        if not field_name:
            continue

        value = field.get("FieldValue", "")

        match_logical = pattern_tariff_logical.match(field_name)
        if match_logical and value:
            reg_id = match_logical.group(1)
            tariff_registers.setdefault(reg_id, {})["hex_obis"] = value.strip()
            continue

        match_scaler = pattern_tariff_scaler.match(field_name)
        if match_scaler and value is not None:
            reg_id = match_scaler.group(1)
            try:
                tariff_registers.setdefault(reg_id, {})["scaler"] = int(value)
            except (TypeError, ValueError):
                pass
            continue

        match_quantity = pattern_tariff_quantity.match(field_name)
        if match_quantity:
            reg_id = match_quantity.group(1)
            tariff_registers.setdefault(reg_id, {})["quantity"] = value.strip()
            continue

    unit_map_from_scaler: Dict[str, str] = {}
    for info in tariff_registers.values():
        hex_obis = info.get("hex_obis")
        quantity = info.get("quantity")
        scaler = info.get("scaler")

        if hex_obis is None or quantity is None or scaler is None:
            continue

        unit = resolve_unit_from_scaler(quantity, scaler)
        if not unit:
            continue

        readable_obis = obis_hex_to_readable(hex_obis)
        unit_map_from_scaler[readable_obis] = unit

    # === 4. Extraire les capture_objects ===
    capture_objects: Dict[int, str] = {}

    # Patterns pour différents formats de fichiers
    # Format 1: DD.Profile_Load01 (LGZ avec 0)
    # Format 2: DD.Profile_Load1 (E360 sans 0)
    pattern_capture_obj_1 = re.compile(
        rf"DD\.Profile_Load{profile_num}\.capture_objects\.0\.(\d+)\.logical_name"
    )
    pattern_capture_obj_2 = re.compile(
        rf"DD\.Profile_Load{int(profile_num)}\.capture_objects\.0\.(\d+)\.logical_name"
    )

    for field in root.iter():
        field_name = field.get("FieldName", "")
        match = pattern_capture_obj_1.match(field_name) or pattern_capture_obj_2.match(field_name)
        if match and field.get("FieldValue"):
            idx = int(match.group(1))
            obis_hex = field.get("FieldValue", "")
            obis_readable = obis_hex_to_readable(obis_hex)
            capture_objects[idx] = obis_readable

    if not capture_objects:
        warnings.append("capture_objects non trouvés dans le fichier XML")
        return []

    # === 5. Initialiser les channels ===
    channels: Dict[str, Dict[str, Any]] = {}

    for idx, obis in capture_objects.items():
        if idx < 2:
            continue
        
        if obis.startswith("0-0:96.") or obis.startswith("0-0:1."):
            continue
        
        # Unité par défaut
        if obis.startswith("1-0:1.") or obis.startswith("1-0:2."):
            unit = "Wh"
        elif obis.startswith("1-0:5.") or obis.startswith("1-0:6.") or \
             obis.startswith("1-0:7.") or obis.startswith("1-0:8."):
            unit = "varh"
        else:
            unit = ""

        detected_unit = unit_map_from_scaler.get(obis, unit)
        channels[obis] = {"unit": detected_unit, "readings": []}

    # === 6. Extraire les données du buffer ===
    # Patterns pour différents formats de fichiers
    # Format LGZ: DD.Profile_Load01.buffer.Selector1.Response.X.Y
    pattern_response_1 = re.compile(
        rf"DD\.Profile_Load{profile_num}\.buffer\.Selector1\.Response\.(\d+)\.(\d+)"
    )
    # Format E360 avec buffer direct: DD.Profile_Load1.buffer.0.X.Y
    pattern_response_2 = re.compile(
        rf"DD\.Profile_Load{int(profile_num)}\.buffer\.0\.(\d+)\.(\d+)"
    )
    # Format LGZ sans 0: (au cas où)
    pattern_response_3 = re.compile(
        rf"DD\.Profile_Load{int(profile_num)}\.buffer\.Selector1\.Response\.(\d+)\.(\d+)"
    )

    response_data: Dict[int, Dict[int, str]] = {}
    timestamp_field_type = ""
    timestamps_are_utc = True  # Par défaut, les payloads XML sont interprétés en UTC
    timezone_offset_checked = False

    for field in root.iter():
        field_name = field.get("FieldName", "")
        match = (pattern_response_1.match(field_name) or
                 pattern_response_2.match(field_name) or
                 pattern_response_3.match(field_name))
        if match:
            row = int(match.group(1))
            col = int(match.group(2))
            value = field.get("FieldValue", "")
            field_type = field.get("FieldType", "")
            
            if col == 0 and not timestamp_field_type and field_type:
                timestamp_field_type = field_type
                # Ne bascule en False que si un offset explicite est trouvé plus loin.
                if "utc" in field_type.lower():
                    timestamps_are_utc = True
            
            if row not in response_data:
                response_data[row] = {}
            response_data[row][col] = value
    
    if not response_data:
        warnings.append("Aucune donnée de mesure trouvée dans le buffer")
        return []

    # === 7. Parser chaque ligne de données ===
    for row in sorted(response_data.keys()):
        row_data = response_data[row]
        
        if 0 not in row_data:
            continue
        
        if (not timezone_offset_checked) and (not timestamp_field_type or "utc" not in timestamp_field_type.lower()):
            hex_ts = row_data[0]
            if isinstance(hex_ts, str) and len(hex_ts) >= 22:
                try:
                    tz_raw = int(hex_ts[18:22], 16)
                    if tz_raw != 0x8000:
                        timestamps_are_utc = False
                except ValueError:
                    pass
            timezone_offset_checked = True
        
        timestamp, dst_value = decode_dlms_timestamp(row_data[0], force_utc=timestamps_are_utc)
        if timestamp is None:
            continue
        
        for col_idx in range(2, max(row_data.keys()) + 1):
            if col_idx not in row_data:
                continue
            
            if col_idx not in capture_objects:
                continue
            
            obis = capture_objects[col_idx]
            if obis not in channels:
                continue
            
            try:
                value = float(row_data[col_idx])
                channels[obis]["readings"].append((timestamp, value, dst_value))
            except (ValueError, TypeError):
                pass

    # === 8. Vérifier qu'on a des données ===
    total_readings = sum(len(ch["readings"]) for ch in channels.values())

    if total_readings == 0:
        warnings.append("Aucune mesure valide extraite du fichier")
        return []

    # === 9. Retourner les données parsées ===
    return [ParsedMeterData(
        meter_id=meter_id,
        load_profile=profile_name,
        interval=interval,
        channels=channels,
        source_file=filename,
        warnings=warnings,
        from_xml=True,  # Flag pour identifier les données XML
        timestamps_utc=timestamps_are_utc
    )]


def parse_xml_billing_values(file_bytes: bytes, filename: str) -> List[ParsedMeterData]:
    """
    Parse un fichier XML DLMS/COSEM de type BillingValues (E570).
    Ces fichiers contiennent des valeurs uniques de registres à un instant T,
    pas des séries temporelles.
    """
    warnings: List[str] = []

    try:
        content = file_bytes.decode('utf-8')
    except UnicodeDecodeError:
        try:
            content = file_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = file_bytes.decode('latin-1')

    try:
        root = ET.fromstring(content)
    except ET.ParseError as e:
        warnings.append(f"Erreur de parsing XML: {str(e)}")
        return []

    # === 1. Extraire DDID (meter ID) ===
    ddid_elem = root.find(".//{http://tempuri.org/DeviceDescriptionDataSet.xsd}DDID")
    if ddid_elem is None:
        ddid_elem = root.find(".//DDID")

    if ddid_elem is None or not ddid_elem.text:
        warnings.append("DDID (meter ID) non trouvé dans le fichier XML")
        return []

    meter_id = ddid_elem.text.strip()

    # === 2. Extraire la date de lecture (ModificationDateTime) ===
    timestamp = None
    mod_datetime_elem = root.find(".//{http://tempuri.org/DeviceDescriptionDataSet.xsd}ModificationDateTime")
    if mod_datetime_elem is None:
        mod_datetime_elem = root.find(".//ModificationDateTime")

    if mod_datetime_elem is not None and mod_datetime_elem.text:
        try:
            # Format: 2025-08-27T12:32:26.7030356+02:00
            timestamp_str = mod_datetime_elem.text.strip()
            # Supprimer les microsecondes excessives (garder 6 chiffres max)
            timestamp_str = re.sub(r'(\.\d{6})\d+', r'\1', timestamp_str)
            timestamp = datetime.fromisoformat(timestamp_str)
        except Exception as e:
            warnings.append(f"Erreur de parsing de la date: {str(e)}")

    # Fallback: utiliser la date actuelle si pas trouvée
    if timestamp is None:
        timestamp = datetime.now(timezone.utc)
        warnings.append("Date de lecture non trouvée, utilisation de la date actuelle")

    # === 3. Extraire tous les registres de type ClassID="3" (Register) ===
    # Pattern pour trouver les objets avec ClassID="3"
    registers: Dict[str, Dict[str, Any]] = {}

    for obj in root.iter():
        if obj.tag.endswith("Objects") or obj.tag == "Objects":
            class_id = obj.get("ClassID")
            if class_id != "3":  # ClassID 3 = Register
                continue

            obj_name = obj.get("ObjectName", "")
            logical_name = obj.get("ObjectLogicalName", "")

            if not logical_name:
                continue

            # Convertir en format lisible OBIS
            obis_readable = obis_hex_to_readable(logical_name)

            # Ne garder que les codes OBIS pertinents (énergie)
            if not (obis_readable.startswith("1-0:1.") or
                    obis_readable.startswith("1-0:2.") or
                    obis_readable.startswith("1-0:3.") or
                    obis_readable.startswith("1-0:4.") or
                    obis_readable.startswith("1-0:5.") or
                    obis_readable.startswith("1-0:6.") or
                    obis_readable.startswith("1-0:7.") or
                    obis_readable.startswith("1-0:8.") or
                    obis_readable.startswith("1-0:9.")):
                continue

            registers[obis_readable] = {
                "obj_name": obj_name,
                "value": None,
                "unit": "",
                "scaler": 0,
                "quantity": ""
            }

    # === 4. Extraire les valeurs et unités des registres ===
    for field in root.iter():
        field_name = field.get("FieldName", "")
        if not field_name:
            continue

        # Patterns pour extraire les informations
        # Format: DD.00003_0100010800FF.CurrentValue
        # Format: DD.00003_0100010800FF.UnitScale.0.Scaler
        # Format: DD.00003_0100010800FF.UnitScale.0.Quantity

        match_current = re.match(r"DD\.[\w_]+\.CurrentValue", field_name)
        if match_current:
            # Extraire le code OBIS du nom du champ
            obis_match = re.search(r"_([0-9A-F]{12})\.", field_name)
            if obis_match:
                obis_hex = obis_match.group(1)
                obis_readable = obis_hex_to_readable(obis_hex)
                if obis_readable in registers:
                    value_str = field.get("FieldValue", "")
                    try:
                        registers[obis_readable]["value"] = float(value_str)
                    except (ValueError, TypeError):
                        pass
            continue

        match_scaler = re.match(r"DD\.[\w_]+\.UnitScale\.0\.Scaler", field_name)
        if match_scaler:
            obis_match = re.search(r"_([0-9A-F]{12})\.", field_name)
            if obis_match:
                obis_hex = obis_match.group(1)
                obis_readable = obis_hex_to_readable(obis_hex)
                if obis_readable in registers:
                    scaler_str = field.get("FieldValue", "")
                    try:
                        registers[obis_readable]["scaler"] = int(scaler_str)
                    except (ValueError, TypeError):
                        pass
            continue

        match_quantity = re.match(r"DD\.[\w_]+\.UnitScale\.0\.Quantity", field_name)
        if match_quantity:
            obis_match = re.search(r"_([0-9A-F]{12})\.", field_name)
            if obis_match:
                obis_hex = obis_match.group(1)
                obis_readable = obis_hex_to_readable(obis_hex)
                if obis_readable in registers:
                    registers[obis_readable]["quantity"] = field.get("FieldValue", "")
            continue

    # === 5. Construire les channels avec une seule lecture ===
    channels: Dict[str, Dict[str, Any]] = {}

    for obis, info in registers.items():
        if info["value"] is None:
            continue

        # Résoudre l'unité
        quantity = info.get("quantity", "")
        scaler = info.get("scaler", 0)
        unit = resolve_unit_from_scaler(quantity, scaler)

        if not unit:
            # Fallback basé sur le code OBIS
            if obis.startswith("1-0:1.") or obis.startswith("1-0:2."):
                unit = "Wh"
            elif obis.startswith("1-0:5.") or obis.startswith("1-0:6."):
                unit = "varh"
            else:
                unit = "Wh"

        # Créer le channel avec une seule mesure (tuple format)
        channels[obis] = {
            "unit": unit,
            "readings": [(timestamp, info["value"], 0)]  # Format: (timestamp, value, dst)
        }

    if not channels:
        warnings.append("Aucun registre valide trouvé dans le fichier BillingValues")
        return []

    # === 6. Retourner les données parsées ===
    return [ParsedMeterData(
        meter_id=meter_id,
        load_profile="Valeurs de facturation",
        interval="NULL",  # Pas d'intervalle pour des valeurs uniques
        channels=channels,
        source_file=filename,
        warnings=warnings,
        from_xml=True,
        timestamps_utc=False  # Timestamp avec timezone explicite
    )]


def parse_xml_flexible(file_bytes: bytes, filename: str, dd_subset: str) -> List[ParsedMeterData]:
    """
    Flexible XML parser pour formats inconnus (fallback layer).
    Découvre automatiquement la structure XML sans assumptions hardcodées.

    Supporte uniquement Load1 et Load2 avec découverte automatique de:
    - Chemins de buffer (Selector1.Response, .0., etc.)
    - Capture objects à profondeur arbitraire
    - Intervalles depuis timestamps réels
    - Unités depuis OBIS codes

    Args:
        file_bytes: Contenu du fichier XML
        filename: Nom du fichier (pour logs)
        dd_subset: Type de données DLMS (ex: "ProfileBuffer")

    Returns:
        Liste de ParsedMeterData (1 par profil détecté)

    Raises:
        XMLParseException: Si la structure ne peut pas être découverte
    """
    warnings = []

    try:
        # Decode bytes to string
        try:
            content = file_bytes.decode('utf-8')
        except UnicodeDecodeError:
            try:
                content = file_bytes.decode('utf-8-sig')
            except UnicodeDecodeError:
                content = file_bytes.decode('latin-1')

        # === 1. Parser XML et créer explorer ===
        root = ET.fromstring(content)
        explorer = XMLStructureExplorer(root)

        # Extraire meter ID (avec namespace fallback)
        meter_id = "UNKNOWN"
        ddid_elem = root.find(".//{http://tempuri.org/DeviceDescriptionDataSet.xsd}DDID")
        if ddid_elem is None:
            ddid_elem = root.find(".//DDID")
        if ddid_elem is not None and ddid_elem.text:
            meter_id = ddid_elem.text

        # === 2. Découvrir structure ===
        structure_map = explorer.discover_structure(meter_id, dd_subset)

        # === 3. Validation: au moins 1 profil trouvé ===
        if not structure_map.profiles:
            analyzed_objects = len(root.findall(".//Object"))
            object_names = [obj.find(".//ObjectName").text for obj in root.findall(".//Object")[:10]
                           if obj.find(".//ObjectName") is not None and obj.find(".//ObjectName").text]

            raise XMLParseException(
                f"Impossible de parser {filename}\n\n"
                f"Raison: Aucun objet profil Load1/Load2 détecté dans le XML\n"
                f"DDSubset: {dd_subset}\n"
                f"Objets analysés: {analyzed_objects}\n"
                f"Profiles trouvés: 0\n\n"
                f"Détails:\n"
                f"- Aucun objet avec ClassID=7 (Profile Generic) et Load1/Load2\n"
                f"- Aucun objet avec attributs 'buffer' ET 'capture_objects'\n"
                f"- ObjectNames examinés: {object_names[:10]}\n\n"
                f"Suggestion: Vérifiez que ce fichier est bien un profil de charge DLMS/COSEM valide.\n"
                f"Si c'est un nouveau format de compteur, contactez le support technique."
            )

        # === 4. Traiter chaque profil découvert ===
        results = []
        resolver = DataPathResolver(root, structure_map)

        for profile_obj in structure_map.profiles:
            profile_name = profile_obj.object_name

            # Validation: profil doit avoir buffer ET capture_objects
            if not profile_obj.has_buffer or not profile_obj.has_capture_objects:
                warnings.append(
                    f"Profil {profile_name} incomplet: buffer={profile_obj.has_buffer}, "
                    f"capture_objects={profile_obj.has_capture_objects}"
                )
                continue

            # Vérifier que capture_objects existe dans la map
            if profile_name not in structure_map.capture_objects:
                warnings.append(f"Pas de capture_objects trouvés pour {profile_name}")
                continue

            capture_objects = structure_map.capture_objects[profile_name]

            # Validation: capture_objects doit avoir au moins timestamp + 1 mesure
            if len(capture_objects) < 2:
                warnings.append(f"{profile_name}: Pas assez de capture_objects ({len(capture_objects)})")
                continue

            # Validation: timestamp doit être en index 0
            if 0 not in capture_objects:
                warnings.append(f"{profile_name}: Pas de timestamp en index 0")
                continue

            # === 4a. Extraire buffer data ===
            try:
                buffer_data = resolver.extract_buffer_data(profile_name)
            except Exception as e:
                warnings.append(f"{profile_name}: Échec extraction buffer - {str(e)}")
                continue

            if not buffer_data:
                warnings.append(f"{profile_name}: Buffer vide")
                continue

            # === 4b. Matcher timestamps aux valeurs ===
            try:
                reading_rows = resolver.match_timestamps_to_values(buffer_data, capture_objects)
            except Exception as e:
                warnings.append(f"{profile_name}: Échec matching timestamps - {str(e)}")
                continue

            if not reading_rows:
                warnings.append(f"{profile_name}: Aucune donnée après matching")
                continue

            # === 4c. Détecter intervalle ===
            detector = ProfileDetector(profile_obj)
            timestamps = [row.timestamp for row in reading_rows]
            interval = detector.detect_interval(timestamps)
            display_name = detector.detect_profile_name(interval)

            # === 4d. Construire channels ===
            channels = {}
            obis_codes = sorted(set(
                obis for row in reading_rows for obis in row.values.keys()
            ))

            for obis in obis_codes:
                # Obtenir unité
                unit = "NULL"
                if obis in structure_map.units:
                    unit = structure_map.units[obis].resolved_unit
                else:
                    # Fallback: inférer depuis OBIS
                    unit = OBISCodeHandler.infer_unit_from_obis(obis, None, None)

                # Construire liste de readings
                readings = []
                for row in reading_rows:
                    if obis in row.values:
                        readings.append({
                            "timestamp": row.timestamp,
                            "value": row.values[obis],
                            "dst": row.dst_value
                        })

                if readings:
                    channels[obis] = {
                        "unit": unit,
                        "readings": readings
                    }

            # Validation finale: au moins 1 channel avec données
            if not channels:
                warnings.append(f"{profile_name}: Aucun channel avec données")
                continue

            # === 4e. Créer ParsedMeterData ===
            results.append(ParsedMeterData(
                meter_id=meter_id,
                load_profile=display_name,
                interval=interval,
                channels=channels,
                source_file=filename,
                warnings=warnings.copy(),
                from_xml=True,
                timestamps_utc=False
            ))

        # === 5. Validation globale: au moins 1 résultat ===
        if not results:
            raise XMLParseException(
                f"Impossible de parser {filename}\n\n"
                f"Raison: Aucun profil valide extrait\n"
                f"DDSubset: {dd_subset}\n"
                f"Profiles détectés: {len(structure_map.profiles)}\n"
                f"Profiles avec données: 0\n\n"
                f"Warnings:\n" + "\n".join(f"- {w}" for w in warnings) + "\n\n"
                f"Suggestion: La structure XML a été détectée mais les données sont incomplètes.\n"
                f"Vérifiez que le fichier contient des mesures valides dans les buffers."
            )

        return results

    except ET.ParseError as e:
        raise XMLParseException(
            f"Impossible de parser {filename}\n\n"
            f"Raison: XML malformé\n"
            f"Erreur: {str(e)}\n\n"
            f"Suggestion: Vérifiez l'intégrité du fichier XML."
        )

    except XMLParseException:
        # Re-raise our custom exceptions
        raise

    except Exception as e:
        raise XMLParseException(
            f"Impossible de parser {filename}\n\n"
            f"Raison: Erreur inattendue pendant la découverte de structure\n"
            f"DDSubset: {dd_subset}\n"
            f"Erreur: {type(e).__name__}: {str(e)}\n\n"
            f"Suggestion: Structure XML non supportée automatiquement.\n"
            f"Contactez le support technique avec ce fichier."
        )


def extract_zip(zip_bytes: bytes) -> List[Tuple[str, bytes]]:
    """Extrait les fichiers d'un ZIP."""
    files: List[Tuple[str, bytes]] = []
    
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes), 'r') as zf:
            for member in zf.namelist():
                if member.endswith('/'):
                    continue
                
                ext = get_file_extension(member)
                if ext in ACCEPTED_DATA_EXTENSIONS:
                    files.append((Path(member).name, zf.read(member)))
    except zipfile.BadZipFile:
        pass
    
    return files


def find_unmapped_obis(data: ParsedMeterData) -> List[str]:
    """Liste les codes OBIS qui n'ont pas de mapping IEC."""
    missing: List[str] = []
    for obis_code, channel in data.channels.items():
        if not channel.get("readings"):
            continue
        if obis_code.startswith(("0-0:96.", "unknown:")):
            continue
        if not OBIS_TO_IEC.get((obis_code, data.interval)) and not OBIS_TO_IEC.get((obis_code, "NULL")):
            missing.append(obis_code)
    return missing


# ============================================================================
# SECTION 5 : CONVERTISSEUR JSON
# ============================================================================

def convert_to_json(data: ParsedMeterData, source: str, mrid_prefix: str) -> Tuple[Dict, str]:
    """
    Convertit ParsedMeterData en structure JSON MeterReadings.
    
    Note: Pour les fichiers XML, le meter_id contient déjà l'identifiant complet
    (ex: LGZ1030767023632), donc on n'ajoute pas le préfixe mRID.
    """
    message_id = str(uuid.uuid4())
    timestamp_now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"
    
    # Pour XML, le meter_id est déjà complet
    if data.from_xml:
        mrid = data.meter_id
    else:
        mrid = f"{mrid_prefix}{data.meter_id}"
    
    # Construire les IntervalBlocks
    interval_blocks: List[Dict] = []
    
    for obis_code, channel_data in data.channels.items():
        if not channel_data["readings"]:
            continue
        
        iec_code = OBIS_TO_IEC.get((obis_code, data.interval))
        if not iec_code:
            iec_code = OBIS_TO_IEC.get((obis_code, "NULL"), "")
        
        if not iec_code:
            continue
        
        unit = channel_data.get("unit", "").lower()
        multiplier = 1000 if unit in UNITS_MULTIPLY_1000 else 1

        for ts, value, dst in channel_data["readings"]:
            formatted_ts = format_timestamp_iso(ts, dst, force_utc=data.timestamps_utc)
            converted_value = int(round(value * multiplier))
            
            interval_blocks.append({
                "IntervalReadings": [
                    {
                        "timeStamp": formatted_ts,
                        "value": str(converted_value),
                        "ReadingQualities": [{"ref": "1.0.0"}]
                    }
                ],
                "ReadingType": {"ref": iec_code}
            })
    
    document = {
        "header": {
            "messageId": message_id,
            "source": source,
            "verb": "created",
            "noun": "MeterReadings",
            "timestamp": timestamp_now
        },
        "payload": {
            "MeterReadings": [{
                "Meter": {
                    "mRID": mrid,
                    "amrSystem": source
                },
                "IntervalBlocks": interval_blocks
            }]
        }
    }
    
    ts_safe = timestamp_now.replace(":", "-")
    filename = f"meter-readings-created_{mrid}_{ts_safe}_{message_id}.json"
    
    return document, filename


# ============================================================================
# SECTION 6 : INTERFACE STREAMLIT
# ============================================================================

def init_session_state():
    """Initialise les variables de session Streamlit."""
    defaults = {
        "processed_data": [],
        "json_outputs": [],
        "warnings": [],
        "mrid_prefix": DEFAULT_MRID_PREFIX,
        "conversion_done": False,
        "files_needing_config": {},
        "user_configs": {}
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val


def render_manual_config_ui(filename: str, preview_lines: List[str], current_data: Optional[ParsedMeterData] = None) -> Optional[Dict]:
    """Affiche l'interface de configuration manuelle si nécessaire."""
    st.warning(f"⚠️ Structure non reconnue automatiquement pour: **{filename}**")
    
    with st.expander("🔧 Configuration manuelle", expanded=True):
        st.markdown("Le fichier n'a pas une structure standard. Veuillez compléter les informations manquantes:")
        
        st.markdown("**Aperçu du fichier:**")
        preview_text = ""
        for i, line in enumerate(preview_lines[:10]):
            preview_text += f"L{i+1}: {line[:100]}{'...' if len(line) > 100 else ''}\n"
        st.code(preview_text, language=None)
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            default_meter = current_data.meter_id if current_data and current_data.meter_id else ""
            meter_id = st.text_input(
                "Numéro du compteur *",
                value=default_meter,
                help="Entrez le numéro de série du compteur (6-12 chiffres)",
                key=f"meter_id_{filename}"
            )
            
            header_line = st.number_input(
                "Ligne des entêtes",
                min_value=1,
                max_value=len(preview_lines),
                value=3,
                help="Numéro de la ligne contenant les noms de colonnes",
                key=f"header_line_{filename}"
            )
        
        with col2:
            data_start = st.number_input(
                "Première ligne de données",
                min_value=1,
                max_value=max(len(preview_lines), 4),
                value=4,
                help="Numéro de la première ligne contenant des mesures",
                key=f"data_start_{filename}"
            )
            
            default_interval_idx = 0
            if current_data and current_data.interval == "24h":
                default_interval_idx = 1
            
            interval = st.selectbox(
                "Intervalle de mesure",
                options=["15min", "24h"],
                index=default_interval_idx,
                help="Fréquence des relevés",
                key=f"interval_{filename}"
            )
        
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            if st.button("✅ Appliquer", type="primary", key=f"apply_{filename}", use_container_width=True):
                if not meter_id or not re.match(r'^\d{6,}$', meter_id):
                    st.error("Le numéro de compteur doit contenir au moins 6 chiffres")
                    return None
                
                return {
                    "meter_id": meter_id,
                    "header_line": int(header_line) - 1,
                    "data_start": int(data_start) - 1,
                    "interval": interval
                }
        
        with col_btn2:
            if st.button("⏭️ Ignorer ce fichier", key=f"skip_{filename}", use_container_width=True):
                return {"skip": True}
    
    return None


def process_uploaded_files(
    files: List,
    delimiter: str,
    source: str,
    mrid_prefix: str
) -> Tuple[List[ParsedMeterData], List[str], Dict]:
    """Traite tous les fichiers uploadés."""
    all_data: List[ParsedMeterData] = []
    all_warnings: List[str] = []
    files_needing_config: Dict = {}
    
    for uploaded_file in files:
        file_bytes = uploaded_file.read()
        ext = get_file_extension(uploaded_file.name)
        
        # === ZIP ===
        if ext == ".zip":
            extracted = extract_zip(file_bytes)
            
            if not extracted:
                all_warnings.append(f"ZIP vide ou invalide: {uploaded_file.name}")
                continue
            
            for inner_name, inner_bytes in extracted:
                inner_ext = get_file_extension(inner_name)
                
                if inner_ext == ".csv":
                    if delimiter == "auto":
                        content = inner_bytes.decode('utf-8', errors='replace')
                        detected_delim = detect_delimiter(content)
                    else:
                        detected_delim = delimiter
                    
                    user_config = st.session_state.user_configs.get(inner_name)
                    results = parse_csv(inner_bytes, inner_name, detected_delim, user_config)
                    
                elif inner_ext in [".xlsx", ".xls"]:
                    results = parse_xlsx(inner_bytes, inner_name)
                    
                elif inner_ext == ".xml":
                    results = parse_xml(inner_bytes, inner_name)
                else:
                    continue
                
                for r in results:
                    if r.needs_user_input:
                        content = inner_bytes.decode('utf-8', errors='replace')
                        files_needing_config[inner_name] = {
                            "lines": content.split('\n'),
                            "data": r,
                            "bytes": inner_bytes
                        }
                    else:
                        all_data.append(r)
                        all_warnings.extend([f"{inner_name}: {w}" for w in r.warnings])
                        
                        missing = find_unmapped_obis(r)
                        if missing:
                            uniq_missing = ", ".join(sorted(set(missing)))
                            all_warnings.append(f"{inner_name}: Codes OBIS non mappés pour {r.interval} (ignorés): {uniq_missing}")
        
        # === CSV ===
        elif ext == ".csv":
            if delimiter == "auto":
                content = file_bytes.decode('utf-8', errors='replace')
                detected_delim = detect_delimiter(content)
            else:
                detected_delim = delimiter
            
            user_config = st.session_state.user_configs.get(uploaded_file.name)
            results = parse_csv(file_bytes, uploaded_file.name, detected_delim, user_config)
            
            for r in results:
                if r.needs_user_input:
                    content = file_bytes.decode('utf-8', errors='replace')
                    files_needing_config[uploaded_file.name] = {
                        "lines": content.split('\n'),
                        "data": r,
                        "bytes": file_bytes
                    }
                else:
                    all_data.append(r)
                    all_warnings.extend([f"{uploaded_file.name}: {w}" for w in r.warnings])
                    
                    missing = find_unmapped_obis(r)
                    if missing:
                        uniq_missing = ", ".join(sorted(set(missing)))
                        all_warnings.append(f"{uploaded_file.name}: Codes OBIS non mappés pour {r.interval} (ignorés): {uniq_missing}")
        
        # === Excel ===
        elif ext in [".xlsx", ".xls"]:
            results = parse_xlsx(file_bytes, uploaded_file.name)
            all_data.extend(results)
            
            for r in results:
                all_warnings.extend([f"{uploaded_file.name}: {w}" for w in r.warnings])
            
            for r in results:
                missing = find_unmapped_obis(r)
                if missing:
                    uniq_missing = ", ".join(sorted(set(missing)))
                    all_warnings.append(f"{uploaded_file.name}: Codes OBIS non mappés pour {r.interval} (ignorés): {uniq_missing}")
        
        # === XML ===
        elif ext == ".xml":
            results = parse_xml(file_bytes, uploaded_file.name)
            all_data.extend(results)
            
            for r in results:
                all_warnings.extend([f"{uploaded_file.name}: {w}" for w in r.warnings])
            
            for r in results:
                missing = find_unmapped_obis(r)
                if missing:
                    uniq_missing = ", ".join(sorted(set(missing)))
                    all_warnings.append(f"{uploaded_file.name}: Codes OBIS non mappés pour {r.interval} (ignorés): {uniq_missing}")
        
        # === Format non supporté ===
        else:
            all_warnings.append(f"Format non supporté: {uploaded_file.name}")
    
    return all_data, all_warnings, files_needing_config


def create_summary_dataframe(
    data_list: List[ParsedMeterData],
    mrid_prefix: str
) -> pd.DataFrame:
    """Crée le tableau récapitulatif des données parsées."""
    rows: List[Dict] = []
    
    for data in data_list:
        # Pour XML, le meter_id est déjà complet
        if data.from_xml:
            mrid = data.meter_id
        else:
            mrid = f"{mrid_prefix}{data.meter_id}"
        
        for obis, ch_data in data.channels.items():
            if not ch_data["readings"]:
                continue
            
            has_mapping = OBIS_TO_IEC.get((obis, data.interval)) or OBIS_TO_IEC.get((obis, "NULL"))
            if not has_mapping:
                continue
            
            timestamps = [r[0] for r in ch_data["readings"]]
            min_ts = min(timestamps)
            max_ts = max(timestamps)
            
            dst_val = ch_data["readings"][0][2]
            offset = resolve_offset(dst_val, data.timestamps_utc)
            
            desc = OBIS_DESCRIPTIONS.get(obis, obis)
            unit = ch_data.get("unit", "")
            
            rows.append({
                "N° Compteur": mrid,
                "Channel": obis,
                "Unité": unit,
                "Date min": min_ts.strftime(f"%Y-%m-%d %H:%M") + offset,
                "Date max": max_ts.strftime(f"%Y-%m-%d %H:%M") + offset,
                "Nb points": len(ch_data["readings"])
            })
    
    return pd.DataFrame(rows)


def create_zip_download(json_outputs: List[Dict]) -> bytes:
    """Crée un fichier ZIP contenant tous les JSON."""
    buffer = io.BytesIO()
    
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for out in json_outputs:
            zf.writestr(out["filename"], out["content"])
    
    buffer.seek(0)
    return buffer.getvalue()


def render_chart(
    data_list: List[ParsedMeterData],
    meter_id: str,
    load_profile: str,
    obis_code: str
):
    """Affiche le graphique de courbe de charge avec Plotly."""
    meter_data = next((d for d in data_list
                       if d.meter_id == meter_id and d.load_profile == load_profile), None)
    
    if not meter_data or obis_code not in meter_data.channels:
        st.warning("Données non trouvées pour cette sélection")
        return
    
    has_mapping = OBIS_TO_IEC.get((obis_code, meter_data.interval)) or OBIS_TO_IEC.get((obis_code, "NULL"))
    if not has_mapping:
        st.warning("Ce canal n'a pas de mapping et n'est pas affiché")
        return
    
    ch_data = meter_data.channels[obis_code]
    
    if not ch_data["readings"]:
        st.warning("Aucune donnée disponible pour ce canal")
        return
    
    unit = ch_data.get("unit", "")
    
    # Pour XML, le meter_id est déjà complet
    if meter_data.from_xml:
        display_mrid = meter_id
    else:
        display_mrid = f"{st.session_state.get('mrid_prefix', DEFAULT_MRID_PREFIX)}{meter_id}"
    
    df = pd.DataFrame([
        {
            "mrID": display_mrid,
            "Channel": obis_code,
            "Unité": unit,
            "Timestamp": r[0],
            "Valeur": r[1]
        }
        for r in ch_data["readings"]
    ]).sort_values("Timestamp")
    
    desc = OBIS_DESCRIPTIONS.get(obis_code, obis_code)
    
    fig = px.line(
        df,
        x="Timestamp",
        y="Valeur",
        title=f"Courbe de charge - {display_mrid} - {desc}",
        labels={"Valeur": f"Valeur ({unit})" if unit else "Valeur", "Timestamp": "Date/Heure"},
        color_discrete_sequence=["#1f77b4"]
    )
    
    fig.update_layout(
        hovermode="x unified",
        xaxis_title="Date/Heure",
        yaxis_title=f"Valeur ({unit})" if unit else "Valeur",
        height=400
    )
    
    fig.update_traces(
        hovertemplate="<b>%{x}</b><br>Valeur: %{y:,.3f}<extra></extra>"
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    with st.expander("Données du graphique", expanded=False):
        st.dataframe(
            df,
            height=300,
            use_container_width=True,
            hide_index=True
        )


def main():
    """Point d'entrée principal de l'application Streamlit."""
    
    st.set_page_config(
        page_title="Convertisseur Relevés Compteurs",
        page_icon="⚡",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    init_session_state()
    
    st.markdown("""
        <style>
        [data-testid="stSidebar"] {
            background-color: #1e3766;
        }
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] .st-emotion-cache-16idsys p,
        [data-testid="stSidebar"] h2 {
            color: #ffffff !important;
        }
        </style>
    """, unsafe_allow_html=True)
    
    # === HEADER ===
    st.title("Convertisseur de relevés de compteurs")
    st.markdown(
        "Création de fichier JSON de **Meter-Readings**."
    )
    
    # === SIDEBAR: CONFIGURATION ===
    with st.sidebar:
        logo_b64 = base64.b64encode(LOGO_SVG.encode('utf-8')).decode('utf-8')
        st.image(f"data:image/svg+xml;base64,{logo_b64}", use_column_width=True)
        
        st.header("Configuration")
        
        source = st.selectbox(
            "Source (AMR System)",
            options=SOURCES,
            index=0,
            help="Système de relevé automatique"
        )
        
        mrid_prefix = st.text_input(
            "Préfixe N° Compteur (.csv, .xlsx, .xls)",
            value=DEFAULT_MRID_PREFIX,
            help="Préfixe ajouté au numéro de compteur pour former le mRID (non utilisé pour les fichiers XML)"
        )
        st.session_state.mrid_prefix = mrid_prefix
        
        delimiter_display = {
            "auto": "Détection auto",
            ";": "Point-virgule (;)",
            ",": "Virgule (,)",
            "\t": "Tabulation (\\t)",
            "|": "Pipe (|)"
        }
        delimiter = st.selectbox(
            "Délimiteur CSV",
            options=list(delimiter_display.keys()),
            format_func=lambda x: delimiter_display[x],
            index=0,
            help="Caractère séparateur des colonnes CSV"
        )
        
        st.markdown("---")

    # === ZONE PRINCIPALE ===
    
    st.header("Import des fichiers")
    
    uploaded_files = st.file_uploader(
        "Glissez-déposez vos fichiers ici",
        type=["csv", "xlsx", "xls", "xml", "zip"],
        accept_multiple_files=True,
        help="Vous pouvez importer plusieurs fichiers simultanément, y compris des archives ZIP"
    )
    
    col1, col2, col3 = st.columns([1, 1, 4])
    
    with col1:
        convert_btn = st.button(
            "Convertir",
            type="primary",
            disabled=not uploaded_files,
            use_container_width=True
        )
    
    with col2:
        if st.session_state.conversion_done:
            clear_btn = st.button(
                "Effacer",
                use_container_width=True
            )
            if clear_btn:
                st.session_state.processed_data = []
                st.session_state.json_outputs = []
                st.session_state.warnings = []
                st.session_state.conversion_done = False
                st.session_state.files_needing_config = {}
                st.session_state.user_configs = {}
                st.rerun()
    
    # === TRAITEMENT ===
    if convert_btn and uploaded_files:
        with st.spinner("Traitement en cours..."):
            data_list, warnings, files_needing_config = process_uploaded_files(
                uploaded_files,
                delimiter,
                source,
                mrid_prefix
            )
            
            st.session_state.processed_data = data_list
            st.session_state.warnings = warnings
            st.session_state.files_needing_config = files_needing_config
            
            json_outputs = []
            for data in data_list:
                doc, fname = convert_to_json(data, source, mrid_prefix)
                
                # Pour XML, le mrid est déjà complet
                if data.from_xml:
                    display_mrid = data.meter_id
                else:
                    display_mrid = f"{mrid_prefix}{data.meter_id}"
                
                json_outputs.append({
                    "filename": fname,
                    "content": json.dumps(doc, indent=4, ensure_ascii=False),
                    "meter_id": data.meter_id,
                    "mrid": display_mrid
                })
            
            st.session_state.json_outputs = json_outputs
            st.session_state.conversion_done = True
        
        if data_list:
            st.success(f"{len(data_list)} compteur(s) traité(s) avec succès!")
        
        if files_needing_config:
            st.info(f"{len(files_needing_config)} fichier(s) nécessitent une configuration manuelle (voir ci-dessous)")
        
        if not data_list and not files_needing_config:
            st.error("Aucun compteur n'a pu être traité. Vérifiez les avertissements.")
    
    # === CONFIGURATION MANUELLE ===
    if st.session_state.files_needing_config:
        st.header("Configuration manuelle requise")
        
        for filename, file_info in list(st.session_state.files_needing_config.items()):
            config = render_manual_config_ui(
                filename, 
                file_info["lines"], 
                file_info.get("data")
            )
            
            if config:
                if config.get("skip"):
                    del st.session_state.files_needing_config[filename]
                    st.rerun()
                else:
                    st.session_state.user_configs[filename] = config
                    
                    file_bytes = file_info["bytes"]
                    content = file_bytes.decode('utf-8', errors='replace')
                    detected_delim = detect_delimiter(content)
                    
                    results = parse_csv(file_bytes, filename, detected_delim, config)
                    
                    if results and not results[0].needs_user_input:
                        st.session_state.processed_data.extend(results)
                        
                        for data in results:
                            doc, fname = convert_to_json(data, source, mrid_prefix)
                            st.session_state.json_outputs.append({
                                "filename": fname,
                                "content": json.dumps(doc, indent=4, ensure_ascii=False),
                                "meter_id": data.meter_id,
                                "mrid": f"{mrid_prefix}{data.meter_id}"
                            })
                        
                        del st.session_state.files_needing_config[filename]
                        st.success(f"✅ {filename} traité avec succès!")
                        st.rerun()
    
    # === AVERTISSEMENTS ===
    if st.session_state.warnings:
        with st.expander(f"⚠️ Avertissements ({len(st.session_state.warnings)})", expanded=False):
            for w in st.session_state.warnings:
                st.write(w)

    st.write("#")

    # === TÉLÉCHARGEMENT ===
    if st.session_state.json_outputs:
        st.header("Téléchargement")
        
        col1, col2 = st.columns(2)
        
        with col1:
            zip_data = create_zip_download(st.session_state.json_outputs)
            timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            st.download_button(
                label="📦 Télécharger tout (ZIP)",
                data=zip_data,
                file_name=f"meter_readings_{timestamp_str}.zip",
                mime="application/zip",
                use_container_width=True
            )
        
        with col2:
            with st.expander("Fichiers individuels", expanded=False):
                for out in st.session_state.json_outputs:
                    display_name = f"{out['mrid']}.json"
                    
                    st.download_button(
                        label=display_name,
                        data=out["content"],
                        file_name=out["filename"],
                        mime="application/json",
                        key=f"dl_{out['filename']}"
                    )

    st.write("#")

    # === VISUALISATION ===
    if st.session_state.processed_data:
        st.header("Visualisation")

        # Créer des clés composites (meter_id, load_profile) pour distinguer les Load Profiles
        meter_keys = [(d.meter_id, d.load_profile) for d in st.session_state.processed_data]

        col1, col2 = st.columns(2)

        with col1:
            # Afficher le bon format selon le type de fichier avec LP si applicable
            def format_meter_display(meter_key):
                meter_id, load_profile = meter_key
                data = next((d for d in st.session_state.processed_data
                            if d.meter_id == meter_id and d.load_profile == load_profile), None)

                # Extraire abréviation LP (LP1, LP2, etc.)
                lp_abbrev = extract_lp_abbreviation(load_profile)
                suffix = f" ({lp_abbrev})" if lp_abbrev else ""

                if data and data.from_xml:
                    return f"{meter_id}{suffix}"
                return f"{st.session_state.mrid_prefix}{meter_id}{suffix}"

            selected_meter_key = st.selectbox(
                "Compteur",
                options=meter_keys,
                format_func=format_meter_display
            )

        # Extraire meter_id et load_profile depuis la clé composite
        selected_meter_id, selected_load_profile = selected_meter_key

        channels: List[Tuple[str, str]] = []
        meter_data = next(
            (d for d in st.session_state.processed_data
             if d.meter_id == selected_meter_id and d.load_profile == selected_load_profile),
            None
        )

        if meter_data:
            for obis in meter_data.channels:
                has_mapping = OBIS_TO_IEC.get((obis, meter_data.interval)) or OBIS_TO_IEC.get((obis, "NULL"))
                if not has_mapping:
                    continue
                unit = meter_data.channels[obis].get("unit", "")
                display = get_obis_display_name(obis, unit)
                channels.append((obis, display))

        with col2:
            if channels:
                selected_idx = st.selectbox(
                    "Canal (Code OBIS)",
                    options=range(len(channels)),
                    format_func=lambda i: channels[i][1]
                )
                selected_obis = channels[selected_idx][0]
            else:
                selected_obis = None
                st.warning("Aucun canal disponible pour ce compteur")
        
        if selected_meter_id and selected_obis:
            render_chart(
                st.session_state.processed_data,
                selected_meter_id,
                selected_load_profile,
                selected_obis
            )

    st.write("#")

    # === RÉCAPITULATIF ===
    if st.session_state.processed_data:
        st.header("Récapitulatif")
        
        summary_df = create_summary_dataframe(
            st.session_state.processed_data,
            st.session_state.mrid_prefix
        )
        
        if not summary_df.empty:
            st.dataframe(
                summary_df,
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("Aucune donnée à afficher")


# ============================================================================
# POINT D'ENTRÉE
# ============================================================================

if __name__ == "__main__":
    main()
